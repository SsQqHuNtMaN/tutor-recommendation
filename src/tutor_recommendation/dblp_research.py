from __future__ import annotations

import hashlib
import json
import os
import re
import subprocess
import time
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

import pandas as pd
import requests
from pypinyin import lazy_pinyin

from .contact_status import STATUS_COLUMN, apply_contact_statuses
from .migrate_contact_status_column import migrate_workbook
from .student_profile import BASE_MATCH_CONTEXT, CORE_TERMS, KEYWORD_WEIGHTS
from .ranking_policy import (
    POLICY_VERSION,
    evaluate_teacher,
    keyword_in_text,
    legacy_dblp_evidence,
    score_text as policy_score_text,
)
from .teacher_identity import TEACHER_ID_COLUMN, ensure_teacher_identity, teacher_record_key
from .cache_utils import configured_max_age_days, read_cached_text
from .run_manifest import context_source_rows, create_run_context, recent_years, write_stage_manifest
from .private_paths import private_file


SCHOOL_SLUG = os.environ.get("SCHOOL_SLUG", "sjtu")
COLLEGE_SLUG = os.environ.get("COLLEGE_SLUG", "cs")
OUTPUT_DIR = Path("outputs") / SCHOOL_SLUG / COLLEGE_SLUG
OUTPUT_PREFIX = f"{SCHOOL_SLUG}_{COLLEGE_SLUG}_teacher_match"
INPUT_PATH = OUTPUT_DIR / f"{OUTPUT_PREFIX}.xlsx"
OUTPUT_PATH = OUTPUT_DIR / f"{OUTPUT_PREFIX}_dblp.xlsx"
CACHE_DIR = OUTPUT_DIR / "dblp_cache"
DBLP_AUTHOR_API = "https://dblp.org/search/author/api"
DBLP_SOURCE_URL = "https://dblp.org/"
TODAY = date.today().isoformat()
RECENT_YEARS = set(recent_years())
AFFILIATION_KEYWORDS = [
    item.strip().lower()
    for item in os.environ.get(
        "AFFILIATION_KEYWORDS",
        "shanghai jiao tong,sjtu,nanjing university,nju",
    ).split(",")
    if item.strip()
]
DBLP_RECOMMENDATION_LEVELS = {
    item.strip()
    for item in os.environ.get("DBLP_RECOMMENDATION_LEVELS", "").split(",")
    if item.strip()
}
DBLP_TARGET_NAMES = {
    item.strip()
    for item in os.environ.get("DBLP_TARGET_NAMES", "").split(",")
    if item.strip()
}
DBLP_SCORE_CAP = 120
DBLP_PRESERVED_COLUMNS = [
    "DBLP匹配状态",
    "DBLP匹配置信度",
    "DBLP作者",
    "DBLP作者链接",
    "DBLP匹配依据",
    "DBLP姓名查询",
    "DBLP近三年论文数",
    "DBLP近三年年份分布",
    "DBLP主要venue",
    "DBLP近三年关键词",
    "DBLP近三年代表论文",
]


def compile_relevant_pattern() -> re.Pattern[str]:
    terms = [keyword for keyword, weight in KEYWORD_WEIGHTS if weight > 0] + list(CORE_TERMS)
    cleaned = [term.strip() for term in terms if term and term.strip()]
    if not cleaned:
        cleaned = ["artificial intelligence", "machine learning", "人工智能", "机器学习"]
    pattern = "|".join(re.escape(term) for term in sorted(set(cleaned), key=len, reverse=True))
    return re.compile(pattern, re.I)


RELEVANT_PATTERN = compile_relevant_pattern()
DBLP_OVERRIDES_PATH = private_file("dblp_overrides.json", env_name="DBLP_OVERRIDES_PATH")


def load_dblp_overrides(
    path: Path = DBLP_OVERRIDES_PATH,
) -> tuple[dict[str, tuple[str, str]], dict[tuple[str, str, str], tuple[str, str]]]:
    if not path.is_file():
        return {}, {}
    data = json.loads(path.read_text(encoding="utf-8"))
    global_overrides: dict[str, tuple[str, str]] = {}
    for name, value in data.get("global", {}).items():
        if isinstance(value, (list, tuple)) and len(value) == 2:
            global_overrides[str(name)] = (str(value[0]), str(value[1]))

    target_overrides: dict[tuple[str, str, str], tuple[str, str]] = {}
    for item in data.get("targets", []):
        if not isinstance(item, dict):
            continue
        key = (
            str(item.get("school_slug", "")).strip(),
            str(item.get("college_slug", "")).strip(),
            str(item.get("name", "")).strip(),
        )
        author = str(item.get("author", "")).strip()
        pid = str(item.get("pid", "")).strip()
        if all(key) and author and pid:
            target_overrides[key] = (author, pid)
    return global_overrides, target_overrides


MANUAL_DBLP_OVERRIDES, TARGET_MANUAL_DBLP_OVERRIDES = load_dblp_overrides()


def norm_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    value = str(value).replace("\u3000", " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", value).strip()


def has_cjk(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", text))


def score_text(text: str) -> tuple[int, list[str]]:
    return policy_score_text(text)


def unique_join(values: list[str], sep: str = "; ") -> str:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        value = norm_text(value)
        if value and value not in seen:
            seen.add(value)
            output.append(value)
    return sep.join(output)


def normalize_author_key(name: str) -> str:
    name = re.sub(r"\s+\d{4}$", "", norm_text(name))
    name = name.lower()
    return re.sub(r"[^a-z]", "", name)


def western_name_candidates(chinese_name: str, teacher_url: str = "") -> list[str]:
    chinese_name = norm_text(chinese_name)
    candidates: list[str] = []

    if re.search(r"[\u4e00-\u9fff]", chinese_name):
        pinyin = lazy_pinyin(chinese_name)
        if len(pinyin) >= 2:
            surname = pinyin[0]
            given = "".join(pinyin[1:])
            given_spaced = " ".join(pinyin[1:])
            candidates.extend(
                [
                    f"{given} {surname}",
                    f"{given_spaced} {surname}",
                    f"{surname} {given}",
                    f"{surname} {given_spaced}",
                ]
            )

    slug = Path(urlparse(norm_text(teacher_url)).path).stem
    slug = re.sub(r"\d+$", "", slug.lower())
    if slug and re.fullmatch(r"[a-z]+", slug):
        pinyin = lazy_pinyin(chinese_name)
        if len(pinyin) >= 2:
            surname = pinyin[0]
            if slug.startswith(surname) and len(slug) > len(surname):
                given = slug[len(surname) :]
                candidates.extend([f"{given} {surname}", f"{surname} {given}"])
            elif slug.endswith(surname) and len(slug) > len(surname):
                given = slug[: -len(surname)]
                candidates.extend([f"{given} {surname}", f"{surname} {given}"])

    return unique_join(candidates).split("; ") if candidates else []


def prepare_url(url: str, params: dict[str, Any] | None = None) -> str:
    if not params:
        return url
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    for key, value in params.items():
        query[key] = [str(value)]
    return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))


def cache_path_for(url: str, suffix: str) -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()
    return CACHE_DIR / f"{digest}.{suffix}"


def valid_payload(text: str, suffix: str) -> bool:
    stripped = text.lstrip()
    lowered = stripped[:300].lower()
    if "gateway time-out" in lowered or lowered.startswith("<html"):
        return False
    if suffix == "json":
        return stripped.startswith("{")
    if suffix == "xml":
        return stripped.startswith("<?xml") or stripped.startswith("<dblpperson")
    return bool(stripped)


def fetch_text(url: str, params: dict[str, Any] | None = None, suffix: str = "txt") -> str:
    full_url = prepare_url(url, params)
    path = cache_path_for(full_url, suffix)
    cached = read_cached_text(path, configured_max_age_days("DBLP_CACHE_MAX_AGE_DAYS", 14))
    if cached is not None:
        if valid_payload(cached, suffix):
            return cached
        path.unlink(missing_ok=True)

    session = requests.Session()
    session.trust_env = False
    headers = {"User-Agent": "Mozilla/5.0 sjtu-teacher-match/1.0"}
    last_error = ""
    for attempt in range(2):
        try:
            response = session.get(full_url, timeout=15, headers=headers)
            if response.status_code == 200 and valid_payload(response.text, suffix):
                path.write_text(response.text, encoding="utf-8")
                return response.text
            last_error = f"HTTP {response.status_code}"
        except Exception as exc:  # noqa: BLE001
            last_error = f"{type(exc).__name__}: {exc}"
        time.sleep(0.5 + attempt)

    if suffix == "xml":
        retry_count = "5"
        retry_delay = "2"
        connect_timeout = "20"
        max_time = "120"
    else:
        retry_count = "2"
        retry_delay = "1"
        connect_timeout = "10"
        max_time = "35"

    curl = [
        "curl.exe",
        "-L",
        "--http1.1",
        "--silent",
        "--show-error",
        "--retry",
        retry_count,
        "--retry-all-errors",
        "--retry-delay",
        retry_delay,
        "--connect-timeout",
        connect_timeout,
        "--max-time",
        max_time,
        "--user-agent",
        "Mozilla/5.0 sjtu-teacher-match/1.0",
        full_url,
    ]
    try:
        completed = subprocess.run(
            curl,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=int(max_time) + 10,
        )
        if completed.returncode == 0 and valid_payload(completed.stdout, suffix):
            path.write_text(completed.stdout, encoding="utf-8")
            return completed.stdout
        last_error = completed.stderr.strip() or f"curl exit {completed.returncode}"
    except Exception as exc:  # noqa: BLE001
        last_error = f"curl {type(exc).__name__}: {exc}"
    raise RuntimeError(last_error)


def as_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def notes_text(info: dict[str, Any]) -> str:
    notes = info.get("notes", {}).get("note") if isinstance(info.get("notes"), dict) else None
    parts: list[str] = []
    for note in as_list(notes):
        if isinstance(note, dict):
            parts.append(norm_text(note.get("text", "")))
        else:
            parts.append(norm_text(note))
    return " ".join(parts)


def extract_dblp_pid(url: str) -> str:
    parsed = urlparse(norm_text(url))
    if "dblp.org" not in parsed.netloc:
        return ""
    path = parsed.path
    path = re.sub(r"\.(html|xml)$", "", path)
    match = re.search(r"/pid/(.+)$", path)
    return match.group(1) if match else ""


def pid_to_url(pid: str) -> str:
    return f"https://dblp.org/pid/{pid}"


def pid_to_xml_url(pid: str) -> str:
    return f"https://dblp.org/pid/{pid}.xml"


def domain_key(url: str) -> str:
    parsed = urlparse(norm_text(url))
    host = parsed.netloc.lower()
    host = host[4:] if host.startswith("www.") else host
    path = parsed.path.rstrip("/").lower()
    return f"{host}{path}"


def homepage_overlap(person_urls: list[str], homepages: list[str]) -> bool:
    person_keys = [domain_key(url) for url in person_urls if domain_key(url)]
    homepage_keys = [domain_key(url) for url in homepages if domain_key(url)]
    for key_a in person_keys:
        for key_b in homepage_keys:
            if key_a and key_b and (key_a in key_b or key_b in key_a):
                return True
    return False


def search_author(query: str) -> list[dict[str, Any]]:
    text = fetch_text(DBLP_AUTHOR_API, {"q": query, "format": "json", "h": 100}, "json")
    data = json.loads(text)
    hits = data.get("result", {}).get("hits", {}).get("hit", [])
    return as_list(hits)


def parse_person_xml(pid: str) -> tuple[list[str], list[dict[str, str]]]:
    text = fetch_text(pid_to_xml_url(pid), suffix="xml")
    root = ET.fromstring(text.encode("utf-8"))
    person_urls = [norm_text(node.text) for node in root.findall("./person/url")]

    pubs: list[dict[str, str]] = []
    for rnode in root.findall("r"):
        if not list(rnode):
            continue
        pub = list(rnode)[0]
        year = norm_text(pub.findtext("year"))
        if year not in RECENT_YEARS:
            continue
        title = norm_text("".join(pub.find("title").itertext()) if pub.find("title") is not None else "")
        venue = norm_text(pub.findtext("booktitle") or pub.findtext("journal") or pub.findtext("series") or "")
        ee = norm_text(pub.findtext("ee") or pub.findtext("url") or "")
        pubs.append(
            {
                "year": year,
                "type": pub.tag,
                "title": title.rstrip("."),
                "venue": venue,
                "ee": ee,
            }
        )
    pubs.sort(key=lambda x: (x["year"], x["venue"], x["title"]), reverse=True)
    return person_urls, pubs


def choose_dblp_author(row: pd.Series) -> dict[str, Any]:
    chinese_name = norm_text(row.get("姓名"))
    teacher_url = norm_text(row.get("教师主页链接"))
    personal_homepage = norm_text(row.get("个人主页"))
    homepages = re.findall(r"https?://[^\s;，,]+", personal_homepage)

    target_key = (SCHOOL_SLUG, COLLEGE_SLUG, chinese_name)
    if target_key in TARGET_MANUAL_DBLP_OVERRIDES:
        author, pid = TARGET_MANUAL_DBLP_OVERRIDES[target_key]
        return {
            "status": "已匹配-高",
            "confidence": "高",
            "author": author,
            "url": pid_to_url(pid),
            "pid": pid,
            "affiliation": "",
            "reason": "目标学院人工DBLP override：已按教师主页/DBLP页面人工确认",
            "score": 200,
            "queries": unique_join(western_name_candidates(chinese_name, teacher_url)),
        }

    if chinese_name in MANUAL_DBLP_OVERRIDES:
        author, pid = MANUAL_DBLP_OVERRIDES[chinese_name]
        return {
            "status": "已匹配-中-旧全局人工覆盖",
            "confidence": "中",
            "author": author,
            "url": pid_to_url(pid),
            "pid": pid,
            "affiliation": "",
            "reason": "旧全局姓名override缺少目标键，仅保留为中置信辅助；请迁移到targets格式",
            "score": 90,
            "queries": unique_join(western_name_candidates(chinese_name, teacher_url)),
        }

    direct_pids: list[str] = []
    for url in homepages:
        pid = extract_dblp_pid(url)
        if pid:
            direct_pids.append(pid)

    queries = western_name_candidates(chinese_name, teacher_url)
    if not queries:
        return {"status": "未匹配", "reason": "无法生成英文姓名候选"}

    query_keys = {normalize_author_key(query) for query in queries}
    candidates: dict[str, dict[str, Any]] = {}

    for pid in direct_pids:
        candidates[pid_to_url(pid)] = {
            "author": "",
            "url": pid_to_url(pid),
            "affiliation": "",
            "direct": True,
        }

    query_errors: list[str] = []
    for query in unique_join(queries).split("; ")[:2]:
        try:
            for hit in search_author(query):
                info = hit.get("info", {})
                url = norm_text(info.get("url"))
                if not url:
                    continue
                candidates[url] = {
                    "author": norm_text(info.get("author")),
                    "url": url,
                    "affiliation": notes_text(info),
                    "direct": False,
                }
        except Exception as exc:  # noqa: BLE001
            query_errors.append(f"{query}: {type(exc).__name__}: {exc}")
            continue

    if not candidates:
        reason = "DBLP作者搜索无候选"
        if query_errors:
            reason += f"；查询错误：{unique_join(query_errors)}"
        return {"status": "未匹配", "reason": reason}

    scored: list[dict[str, Any]] = []
    for candidate in candidates.values():
        author = candidate["author"]
        candidate_key = normalize_author_key(author)
        affiliation = candidate["affiliation"].lower()
        score = 0
        reasons: list[str] = []
        if candidate.get("direct"):
            score += 120
            reasons.append("教师主页含DBLP链接")
        if candidate_key in query_keys:
            score += 65
            reasons.append("姓名精确匹配")
        elif any(key and (key in candidate_key or candidate_key in key) for key in query_keys):
            score += 25
            reasons.append("姓名近似匹配")
        matched_affiliation = [keyword for keyword in AFFILIATION_KEYWORDS if keyword in affiliation]
        if matched_affiliation:
            score += 90
            reasons.append(f"DBLP affiliation含{matched_affiliation[0]}")
        elif "shanghai" in affiliation or "nanjing" in affiliation:
            score += 25
            reasons.append("DBLP affiliation含目标城市线索")

        candidate["score"] = score
        candidate["score_reasons"] = reasons
        candidate["name_match"] = "exact" if candidate_key in query_keys else "approx" if any(
            key and (key in candidate_key or candidate_key in key) for key in query_keys
        ) else "none"
        candidate["affiliation_match"] = bool(matched_affiliation)
        candidate["homepage_match"] = False
        scored.append(candidate)

    scored.sort(key=lambda item: item["score"], reverse=True)

    # Use homepage overlap only when it can materially disambiguate; XML requests are the slow part.
    if homepages and scored and scored[0]["score"] < 150:
        for candidate in [item for item in scored[:3] if item["score"] >= 50]:
            pid = extract_dblp_pid(candidate["url"])
            if not pid:
                continue
            try:
                person_urls, _ = parse_person_xml(pid)
            except Exception:
                person_urls = []
            if homepage_overlap(person_urls, homepages):
                candidate["score"] += 100
                candidate["score_reasons"].append("DBLP个人URL与教师主页重合")
                candidate["homepage_match"] = True

    scored.sort(key=lambda item: item["score"], reverse=True)
    best = scored[0]
    runner_up_score = scored[1]["score"] if len(scored) > 1 else 0

    margin = best["score"] - runner_up_score
    if best.get("direct") or best.get("homepage_match"):
        confidence = "高"
    elif best.get("name_match") == "exact" and best.get("affiliation_match") and margin >= 20:
        confidence = "高"
    elif best.get("name_match") == "exact" and margin >= 20:
        confidence = "中"
    elif best.get("name_match") == "approx" and best.get("affiliation_match") and margin >= 30:
        confidence = "低"
    else:
        confidence = "未采用"

    if confidence == "未采用":
        best_names = " | ".join(
            f"{item.get('author') or item.get('url')}({item.get('score')})" for item in scored[:5]
        )
        return {
            "status": "歧义/未采用",
            "reason": f"候选不足以可靠消歧：{best_names}",
            "queries": unique_join(queries),
        }

    pid = extract_dblp_pid(best["url"])
    author_name = best["author"]
    if not author_name and pid:
        try:
            text = fetch_text(pid_to_xml_url(pid), suffix="xml")
            root = ET.fromstring(text.encode("utf-8"))
            author_node = root.find("./person/author")
            author_name = norm_text(author_node.text if author_node is not None else "")
        except Exception:
            author_name = ""

    return {
        "status": f"已匹配-{confidence}",
        "confidence": confidence,
        "author": author_name,
        "url": pid_to_url(pid) if pid else best["url"],
        "pid": pid,
        "affiliation": best["affiliation"],
        "reason": unique_join(best["score_reasons"]),
        "score": best["score"],
        "queries": unique_join(queries),
    }


def summarize_pubs(pubs: list[dict[str, str]]) -> dict[str, Any]:
    if not pubs:
        return {
            "count": 0,
            "year_counts": "",
            "venues": "",
            "titles": "",
            "keywords": "",
            "dblp_score": 0,
            "has_core": False,
        }

    title_text = " ".join(f"{pub['title']} {pub['venue']}" for pub in pubs)
    dblp_score, matched = score_text(title_text)
    counts = Counter(pub["year"] for pub in pubs)
    venues = Counter(pub["venue"] for pub in pubs if pub["venue"])

    def relevance(pub: dict[str, str]) -> tuple[int, str]:
        score, _ = score_text(f"{pub['title']} {pub['venue']}")
        return score, pub["year"]

    representative = sorted(pubs, key=relevance, reverse=True)[:10]
    title_items = [
        f"[{pub['year']} {pub['venue'] or pub['type']}] {pub['title']}"
        for pub in representative
    ]
    has_core = any(keyword in CORE_TERMS for keyword in matched)
    return {
        "count": len(pubs),
        "year_counts": "; ".join(f"{year}:{counts[year]}" for year in sorted(counts, reverse=True)),
        "venues": "; ".join(f"{venue}({count})" for venue, count in venues.most_common(10)),
        "titles": "；".join(title_items),
        "keywords": unique_join(matched),
        "dblp_score": dblp_score,
        "has_core": has_core,
    }


def update_recommendation(row: pd.Series, dblp_summary: dict[str, Any], match: dict[str, Any]) -> dict[str, Any]:
    decision = evaluate_teacher(row, dblp=legacy_dblp_evidence(row, dblp_summary, match))
    update = decision.to_columns()
    direction = norm_text(row.get("研究方向"))
    dblp_direction = norm_text(dblp_summary.get("keywords"))
    update["综合研究方向（主页+DBLP）"] = unique_join(
        [direction, f"DBLP近三年关键词：{dblp_direction}" if dblp_direction else ""]
    )[:1200]
    return update


def should_query_dblp(row: pd.Series) -> bool:
    if norm_text(row.get("是否建议套磁")) != "是":
        return False
    if DBLP_RECOMMENDATION_LEVELS and norm_text(row.get("推荐等级")) not in DBLP_RECOMMENDATION_LEVELS:
        return False
    if DBLP_TARGET_NAMES and norm_text(row.get("姓名")) not in DBLP_TARGET_NAMES:
        return False
    return True


def teacher_row_key(row: pd.Series | dict[str, Any]) -> str:
    return teacher_record_key(SCHOOL_SLUG, COLLEGE_SLUG, row)


def load_existing_dblp_output() -> tuple[dict[str, dict[str, Any]], pd.DataFrame]:
    if not OUTPUT_PATH.exists():
        return {}, pd.DataFrame()
    try:
        full_df = pd.read_excel(OUTPUT_PATH, sheet_name="全量教师名录")
    except Exception:  # noqa: BLE001
        full_df = pd.DataFrame()
    previous_rows = {
        teacher_row_key(row): row.to_dict()
        for _, row in full_df.iterrows()
        if norm_text(row.get("姓名"))
    }
    try:
        pubs_df = pd.read_excel(OUTPUT_PATH, sheet_name="DBLP近三年论文明细")
    except Exception:  # noqa: BLE001
        pubs_df = pd.DataFrame()
    return previous_rows, pubs_df


def merge_current_row_with_previous_dblp(row: pd.Series, previous: dict[str, Any]) -> dict[str, Any]:
    merged = row.to_dict()
    for col in DBLP_PRESERVED_COLUMNS:
        if col in previous:
            merged[col] = previous.get(col)
    previous_combined = norm_text(previous.get("综合研究方向（主页+DBLP）"))
    if previous_combined:
        merged["综合研究方向（主页+DBLP）"] = unique_join(
            [
                row.get("研究方向"),
                previous.get("DBLP近三年关键词"),
                previous.get("DBLP近三年代表论文"),
            ]
        )[:1200] or previous_combined
    else:
        merged["综合研究方向（主页+DBLP）"] = row.get("研究方向")
    return merged


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill("solid", fgColor="D9EAD3")
    mid_fill = PatternFill("solid", fgColor="FFF2CC")
    low_fill = PatternFill("solid", fgColor="F4CCCC")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(idx)
            header = norm_text(ws.cell(row=1, column=idx).value)
            if header in {"推荐理由", "研究方向", "综合研究方向（主页+DBLP）", "DBLP近三年代表论文"}:
                ws.column_dimensions[col_letter].width = 58
            elif header in {"教师主页链接", "DBLP作者链接", "个人主页"}:
                ws.column_dimensions[col_letter].width = 38
            elif header in {"命中关键词", "DBLP近三年关键词", "DBLP主要venue"}:
                ws.column_dimensions[col_letter].width = 30
            else:
                ws.column_dimensions[col_letter].width = 16

        headers = [cell.value for cell in ws[1]]
        if ws.max_row >= 2 and "推荐等级" in headers:
            level_col = headers.index("推荐等级") + 1
            col_letter = get_column_letter(level_col)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"强烈建议"'], fill=high_fill),
            )
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"可以考虑"'], fill=mid_fill),
            )
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"暂不优先"'], fill=low_fill),
            )

    wb.save(path)


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"未找到输入表：{INPUT_PATH}")

    run_context = create_run_context("dblp", f"{SCHOOL_SLUG}_{COLLEGE_SLUG}", [INPUT_PATH])
    df = pd.read_excel(INPUT_PATH, sheet_name="全量教师名录")
    df = pd.DataFrame(
        ensure_teacher_identity(SCHOOL_SLUG, COLLEGE_SLUG, row.to_dict())
        for _, row in df.iterrows()
    )
    df = apply_contact_statuses(df, SCHOOL_SLUG, COLLEGE_SLUG)
    previous_rows, previous_pubs = load_existing_dblp_output()
    query_keys = {
        teacher_row_key(row)
        for _, row in df.iterrows()
        if should_query_dblp(row)
    }
    output_rows: list[dict[str, Any]] = []
    publication_rows: list[dict[str, Any]] = []
    if not previous_pubs.empty:
        for _, pub_row in previous_pubs.iterrows():
            if teacher_row_key(pub_row) not in query_keys:
                publication_rows.append(pub_row.to_dict())

    for idx, row in df.iterrows():
        if (idx + 1) % 10 == 0:
            print(f"processing {idx + 1}/{len(df)}...", flush=True)

        row_dict = row.to_dict()
        pubs: list[dict[str, str]] = []
        if should_query_dblp(row):
            print(f"querying DBLP {row.get('姓名')}", flush=True)
            match = choose_dblp_author(row)
            if match.get("pid") and match.get("confidence") in {"高", "中"}:
                try:
                    _, pubs = parse_person_xml(match["pid"])
                except Exception as exc:  # noqa: BLE001
                    match["status"] = "DBLP论文抓取失败"
                    match["reason"] = f"{type(exc).__name__}: {exc}"
            summary = summarize_pubs(pubs)
            rec_update = update_recommendation(row, summary, match)
        else:
            previous = previous_rows.get(teacher_row_key(row))
            if previous:
                output_rows.append(merge_current_row_with_previous_dblp(row, previous))
                continue
            match = {"status": "未检索-非当前套磁名单", "reason": "原表未标注为可套磁，未纳入本轮DBLP优先检索"}
            summary = summarize_pubs([])
            rec_update = {
                "推荐等级": row.get("推荐等级"),
                "是否建议套磁": row.get("是否建议套磁"),
                "匹配分": row.get("匹配分"),
                "命中关键词": row.get("命中关键词"),
                "综合研究方向（主页+DBLP）": row.get("研究方向"),
                "推荐理由": row.get("推荐理由"),
            }

        row_dict.update(rec_update)
        row_dict.update(
            {
                "DBLP匹配状态": match.get("status", ""),
                "DBLP匹配置信度": match.get("confidence", ""),
                "DBLP作者": match.get("author", ""),
                "DBLP作者链接": match.get("url", ""),
                "DBLP匹配依据": match.get("reason", ""),
                "DBLP姓名查询": match.get("queries", ""),
                "评分规则版本": POLICY_VERSION,
                "DBLP近三年论文数": summary["count"],
                "DBLP近三年年份分布": summary["year_counts"],
                "DBLP主要venue": summary["venues"],
                "DBLP近三年关键词": summary["keywords"],
                "DBLP近三年代表论文": summary["titles"],
            }
        )
        output_rows.append(row_dict)

        for pub in pubs:
            publication_rows.append(
                {
                    "姓名": row.get("姓名"),
                    TEACHER_ID_COLUMN: row.get(TEACHER_ID_COLUMN, ""),
                    STATUS_COLUMN: row.get(STATUS_COLUMN, ""),
                    "DBLP作者": match.get("author", ""),
                    "年份": pub["year"],
                    "类型": pub["type"],
                    "venue": pub["venue"],
                    "题名": pub["title"],
                    "链接": pub["ee"],
                    "教师主页链接": row.get("教师主页链接"),
                    "DBLP作者链接": match.get("url", ""),
                }
            )

    out_df = pd.DataFrame(output_rows)
    out_df = apply_contact_statuses(out_df, SCHOOL_SLUG, COLLEGE_SLUG)
    pubs_df = pd.DataFrame(publication_rows)
    if not pubs_df.empty:
        dedup_columns = [
            column
            for column in [TEACHER_ID_COLUMN, "DBLP作者链接", "年份", "题名", "链接"]
            if column in pubs_df.columns
        ]
        pubs_df = pubs_df.drop_duplicates(subset=dedup_columns, keep="last")
        pubs_df = apply_contact_statuses(pubs_df, SCHOOL_SLUG, COLLEGE_SLUG)
    priority_order = {"强烈建议": 0, "可以考虑": 1, "暂不优先": 2}
    priority_df = out_df[out_df["是否建议套磁"] == "是"].copy()
    priority_df["_order"] = priority_df["推荐等级"].map(priority_order).fillna(9)
    priority_df = priority_df.sort_values(["_order", "匹配分", "名录序号"], ascending=[True, False, True])
    priority_df = priority_df.drop(columns=["_order"])

    full_df = out_df.copy()
    full_df["_order"] = full_df["推荐等级"].map(priority_order).fillna(9)
    full_df = full_df.sort_values(["_order", "匹配分", "名录序号"], ascending=[True, False, True])
    full_df = full_df.drop(columns=["_order"])

    source_df = pd.DataFrame(
        [
            {"项目": "更新日期", "内容": TODAY},
            {"项目": "DBLP来源", "内容": DBLP_SOURCE_URL},
            {"项目": "近三年口径", "内容": "按DBLP年份字段取" + "、".join(sorted(RECENT_YEARS))},
            {"项目": "匹配依据", "内容": BASE_MATCH_CONTEXT},
            {
                "项目": "DBLP作者消歧说明",
                "内容": f"优先使用姓名精确匹配、DBLP affiliation含目标学校关键词（{', '.join(AFFILIATION_KEYWORDS)}）、教师主页与DBLP个人URL重合。低置信度和歧义项在表中保留状态，低置信度论文仅作参考。",
            },
            {
                "项目": "DBLP检索范围",
                "内容": "本轮为避免DBLP限流，优先对上一版已标注为可套磁的教师做DBLP作者匹配和近三年论文抓取；其余教师保留全量信息，并在DBLP匹配状态中标注“未检索-非当前套磁名单”。",
            },
        ]
        + context_source_rows(run_context)
    )
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        priority_df.to_excel(writer, sheet_name="优先套磁名单", index=False)
        full_df.to_excel(writer, sheet_name="全量教师名录", index=False)
        pubs_df.to_excel(writer, sheet_name="DBLP近三年论文明细", index=False)
        source_df.to_excel(writer, sheet_name="匹配依据", index=False)

    style_workbook(OUTPUT_PATH)
    migrate_workbook(OUTPUT_PATH, {})
    write_stage_manifest(OUTPUT_DIR, run_context)

    print(f"rows={len(full_df)}", flush=True)
    print(f"publications={len(pubs_df)}", flush=True)
    print(f"output={OUTPUT_PATH.resolve()}", flush=True)
    print(full_df["推荐等级"].value_counts().to_string(), flush=True)
    print(full_df["DBLP匹配状态"].value_counts().to_string(), flush=True)
    print("\nTop recommendations:", flush=True)
    print(
        priority_df[
            [
                "姓名",
                "职称",
                "推荐等级",
                "匹配分",
                "DBLP近三年论文数",
                "DBLP近三年关键词",
                "DBLP作者链接",
                "教师主页链接",
            ]
        ]
        .head(25)
        .to_string(index=False)
        ,
        flush=True,
    )


if __name__ == "__main__":
    main()
