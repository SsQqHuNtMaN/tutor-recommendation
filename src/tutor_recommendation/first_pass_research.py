from __future__ import annotations

import argparse
import hashlib
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import pandas as pd
import requests
from bs4 import BeautifulSoup

from .contact_status import apply_contact_statuses
from .migrate_contact_status_column import migrate_workbook
from .student_profile import (
    HIGH_SIGNAL_TERMS,
    INSTITUTE_BONUS,
    KEYWORD_WEIGHTS,
    RESUME_MATCH_CONTEXT,
)
from .teacher_match_targets import TARGETS, TargetConfig
from .teacher_identity import (
    IDENTITY_CONFIDENCE_COLUMN,
    TEACHER_ID_COLUMN,
    ensure_teacher_identity,
)
from .ranking_policy import evaluate_teacher, keyword_in_text as policy_keyword_in_text
from .run_manifest import context_source_rows, create_run_context, write_stage_manifest


TODAY = date.today().isoformat()
SJTU_CS_BASE_URL = "https://www.cs.sjtu.edu.cn"
SJTU_CS_LIST_API_URL = f"{SJTU_CS_BASE_URL}/active/ajax_teacher_list.html"
ZJU_CS_TEAM_INFO_URL = "http://www.cs.zju.edu.cn/csen/2021/0525/c27006a2377953/page.htm"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0 Safari/537.36 faculty-match/1.0"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}
ILLEGAL_EXCEL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def has_cjk(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", text))


def keyword_in_text(keyword: str, text_lower: str) -> bool:
    return policy_keyword_in_text(keyword, text_lower)


def norm_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    value = str(value).replace("\u3000", " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", value).strip()


def clean_name(name: str) -> str:
    name = norm_text(name)
    name = re.sub(r"[（(].*?[）)]", "", name).strip()
    if re.fullmatch(r"[\u4e00-\u9fff ]{2,8}", name):
        name = name.replace(" ", "")
    return name


def split_label_value(text: str) -> tuple[str, str]:
    text = norm_text(text)
    if "：" in text:
        label, value = text.split("：", 1)
    elif ":" in text:
        label, value = text.split(":", 1)
    else:
        return "", text
    return label.replace(" ", ""), norm_text(value)


def unique_join(values: list[str] | tuple[str, ...], sep: str = "; ") -> str:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        value = norm_text(value)
        if value and value not in seen:
            seen.add(value)
            output.append(value)
    return sep.join(output)


def normalize_identity_url(value: Any) -> str:
    text = norm_text(value)
    if not text:
        return ""
    text = text.split(";")[0].strip()
    parsed = urlparse(text)
    if not parsed.scheme or not parsed.netloc:
        return text.rstrip("/#")
    host = parsed.netloc.lower()
    path = parsed.path.rstrip("/") or "/"
    query = urlencode(sorted(parse_qsl(parsed.query, keep_blank_values=True)))
    return urlunparse((parsed.scheme.lower(), host, path, "", query, ""))


def is_generic_identity_url(value: str) -> bool:
    url = normalize_identity_url(value)
    if not url:
        return True
    if url in GENERIC_IDENTITY_URLS:
        return True
    parsed = urlparse(url)
    host = parsed.netloc.lower()
    path = parsed.path.rstrip("/")
    if host in GENERIC_NON_HOMEPAGE_HOSTS:
        return True
    if host == "aiia.seu.edu.cn":
        return True
    if parsed.fragment:
        return True
    if host in GENERIC_COMMON_HOMEPAGE_HOSTS and path in {"", "/", "/index.htm", "/index.html"}:
        return True
    if host in GENERIC_COMMON_HOMEPAGE_HOSTS and path.endswith("/list.htm"):
        return True
    if host in {"www.cad.zju.edu.cn", "cad.zju.edu.cn"} and path in {"", "/", "/english.html"}:
        return True
    if host in {"cs.seu.edu.cn", "cse.seu.edu.cn"} and path in {"", "/", "/dsxx/list.htm", "/54820/list.htm"}:
        return True
    return False


def row_identity_keys(row: dict[str, Any]) -> list[str]:
    keys: list[str] = []
    for column in ["个人主页", "教师主页链接"]:
        raw_url = norm_text(row.get(column))
        if "#" in raw_url:
            continue
        url = normalize_identity_url(raw_url)
        if not url or is_generic_identity_url(url):
            continue
        key = f"url:{url}"
        if key not in keys:
            keys.append(key)
    return keys


def row_cross_target_name_key(row: dict[str, Any], config: TargetConfig) -> str:
    name = clean_name(norm_text(row.get("姓名")))
    if not re.fullmatch(r"[\u4e00-\u9fff]{2,8}", name):
        return ""
    return f"name:{config.school_slug}:{name}"


MERGE_DUPLICATE_COLUMNS = [
    "职称",
    "名录研究所",
    "主页研究所",
    "官方系别",
    "所内职务",
    "是否兼职",
    "邮箱",
    "电话",
    "地址",
    "个人主页",
    "教师主页链接",
    "研究方向",
    "导师信息库研究方向",
    "导师信息库团队",
    "团队PDF证据",
    "导师信息库PDF",
    "个人简介摘要",
]

GENERIC_IDENTITY_URLS = {
    "https://chenyang10.github.io/chengyang",
    "https://chenyang10.github.io/chengyang/",
    "https://manage.faculty.ustc.edu.cn/system/caslogin.jsp",
    "http://sc.ruc.edu.cn/department/ssai/ssai_users/index.htm",
    "http://www.fudanmagiclab.com",
    "http://www.fudanmagiclab.com/",
    "https://dml.fudan.edu.cn",
    "https://dml.fudan.edu.cn/",
}


def merge_duplicate_row(existing: dict[str, Any], duplicate: dict[str, Any], reason: str) -> dict[str, Any]:
    merged = dict(existing)
    for column in MERGE_DUPLICATE_COLUMNS:
        merged[column] = unique_join([merged.get(column, ""), duplicate.get(column, "")])
    merged["去重备注"] = unique_join([merged.get("去重备注", ""), reason])
    return merged


def deduplicate_faculty_rows(rows: list[dict[str, Any]], config: TargetConfig) -> list[dict[str, Any]]:
    seen_key_to_index: dict[str, int] = {}
    output: list[dict[str, Any]] = []
    skipped = 0
    for row in rows:
        keys = row_identity_keys(row)
        duplicate_index = next((seen_key_to_index[key] for key in keys if key in seen_key_to_index), None)
        if duplicate_index is not None:
            skipped += 1
            kept = output[duplicate_index]
            reason = f"同目标重复记录合并：{norm_text(row.get('姓名'))} @ {config.key}"
            output[duplicate_index] = merge_duplicate_row(kept, row, reason)
            for key in keys:
                seen_key_to_index.setdefault(key, duplicate_index)
            continue
        output.append(dict(row))
        new_index = len(output) - 1
        for key in keys:
            seen_key_to_index.setdefault(key, new_index)
    if skipped:
        print(f"{config.key}_dedup_within={skipped}")
    for index, row in enumerate(output, start=1):
        row["名录序号"] = index
    return output


def deduplicate_targets_rows(
    rows_by_target: dict[str, list[dict[str, Any]]],
    configs: list[TargetConfig],
) -> dict[str, list[dict[str, Any]]]:
    if len(configs) <= 1:
        return rows_by_target

    ordered_configs = sorted(configs, key=lambda item: (item.dedup_priority, list(TARGETS).index(item.key)))
    seen_key_to_owner: dict[str, tuple[str, int]] = {}
    seen_name_to_owner: dict[str, tuple[str, int]] = {}
    kept: dict[str, list[dict[str, Any]]] = {config.key: [] for config in configs}
    skipped_by_target: dict[str, int] = {config.key: 0 for config in configs}

    for config in ordered_configs:
        for row in rows_by_target.get(config.key, []):
            keys = row_identity_keys(row)
            name_key = row_cross_target_name_key(row, config)
            duplicate_key = next((key for key in keys if key in seen_key_to_owner), "")
            duplicate_owner = seen_key_to_owner.get(duplicate_key) if duplicate_key else None
            duplicate_reason = "identity"
            if not duplicate_owner and name_key and name_key in seen_name_to_owner:
                candidate_owner = seen_name_to_owner[name_key]
                candidate_row = kept[candidate_owner[0]][candidate_owner[1]]
                duplicate_name = norm_text(row.get("姓名"))
                review_note = f"同校同名待人工复核：{duplicate_name} @ {config.key} / {candidate_owner[0]}"
                candidate_row["去重备注"] = unique_join([candidate_row.get("去重备注", ""), review_note])
                row["去重备注"] = unique_join([row.get("去重备注", ""), review_note])
            if duplicate_owner:
                skipped_by_target[config.key] += 1
                owner_target, owner_index = duplicate_owner
                owner_row = kept[owner_target][owner_index]
                duplicate_name = norm_text(row.get("姓名"))
                reason = (
                    f"跨目标重复记录合并（{duplicate_reason}）："
                    f"{duplicate_name} @ {config.key} -> {owner_target}"
                )
                kept[owner_target][owner_index] = merge_duplicate_row(owner_row, row, reason)
                for key in keys:
                    seen_key_to_owner.setdefault(key, duplicate_owner)
                if name_key:
                    seen_name_to_owner.setdefault(name_key, duplicate_owner)
                continue
            kept[config.key].append(row)
            owner = (config.key, len(kept[config.key]) - 1)
            for key in keys:
                seen_key_to_owner.setdefault(key, owner)
            if name_key:
                seen_name_to_owner.setdefault(name_key, owner)

    total_skipped = sum(skipped_by_target.values())
    if total_skipped:
        summary = ", ".join(f"{key}:{count}" for key, count in skipped_by_target.items() if count)
        print(f"global_dedup_skipped={total_skipped} ({summary})")
    for config in configs:
        for index, row in enumerate(kept[config.key], start=1):
            row["名录序号"] = index
    return kept


def score_teacher(row: dict[str, Any]) -> dict[str, Any]:
    return evaluate_teacher(row).to_columns()


NJU_ACADEMIC_CATEGORIES = [
    ("教授", "https://cs.nju.edu.cn/2639/list.htm"),
    ("副教授", "https://cs.nju.edu.cn/2640/list.htm"),
    ("准长聘", "https://cs.nju.edu.cn/zzp/list.htm"),
    ("跨学科博导", "https://cs.nju.edu.cn/kxkbd/list.htm"),
    ("讲师、专职科研、博士后", "https://cs.nju.edu.cn/2641/list.htm"),
    ("高级工程师", "https://cs.nju.edu.cn/2642/list.htm"),
    ("专业技术人员", "https://cs.nju.edu.cn/2643/list.htm"),
]

NJU_AI_ACADEMIC_SECTIONS = {"教师", "专职科研、博士后"}


RUC_GSAI_NAME_OVERRIDES = {
    "XiaoZHOU": "周骁",
}

RUC_COMMON_HOSTS = {
    "ai.ruc.edu.cn",
    "gsai.ruc.edu.cn",
    "www.ruc.edu.cn",
    "cmst.ruc.edu.cn",
}

RUC_NON_HOMEPAGE_HOSTS = {
    "academic.oup.com",
    "arxiv.org",
    "dblp.uni-trier.de",
    "dl.acm.org",
    "doi.org",
    "ieeexplore.ieee.org",
    "doi.ieeecomputersociety.org",
    "link.springer.com",
    "nature.com",
    "www.nature.com",
    "openreview.net",
    "proceedings.neurips.cc",
    "scholar.google.com",
}

NJU_IS_ACADEMIC_TITLES = {
    "教授",
    "副教授",
    "长聘副教授",
    "准聘副教授",
    "准聘助理教授",
    "兼职教授",
    "专职科研",
    "博士后",
}

FUDAN_CIRAM_ADMIN_UNITS = {
    "党政办公室",
    "科研学科办公室",
    "教学管理办公室",
    "学生工作办公室",
    "党政办公室/科研学科办公室",
}

GENERIC_COMMON_HOMEPAGE_HOSTS = {
    "www.nju.edu.cn",
    "ic.nju.edu.cn",
    "njusz.nju.edu.cn",
    "www.ruc.edu.cn",
    "sc.ruc.edu.cn",
    "www.fudan.edu.cn",
    "ai.fudan.edu.cn",
    "ciram.fudan.edu.cn",
    "cse.seu.edu.cn",
    "aiia.seu.edu.cn",
    "ai.seu.edu.cn",
    "cs.seu.edu.cn",
    "jssec.seu.edu.cn",
    "palm.seu.edu.cn",
    "cs.tongji.edu.cn",
    "see.tongji.edu.cn",
    "www.tongji.edu.cn",
    "www.zju.edu.cn",
    "www.cs.zju.edu.cn",
    "www.cse.zju.edu.cn",
    "ai.zju.edu.cn",
    "person.zju.edu.cn",
    "mypage.zju.edu.cn",
    "zjui.intl.zju.edu.cn",
    "www.ustc.edu.cn",
    "faculty.ustc.edu.cn",
    "faculty-en.ustc.edu.cn",
    "manage.faculty.ustc.edu.cn",
    "www.lamda.nju.edu.cn",
    "www.fudanmagiclab.com",
    "dml.fudan.edu.cn",
}

GENERIC_NON_HOMEPAGE_HOSTS = {
    "scholar.google.com",
    "dblp.org",
    "dblp.uni-trier.de",
    "doi.org",
    "dl.acm.org",
    "ieeexplore.ieee.org",
    "link.springer.com",
    "orcid.org",
    "www.orcid.org",
    "researchgate.net",
    "www.researchgate.net",
}


def fetch(session: requests.Session, url: str, encoding: str = "utf-8") -> BeautifulSoup:
    response = session.get(url, timeout=30, headers=HEADERS)
    response.raise_for_status()
    response.encoding = encoding
    return BeautifulSoup(response.text, "html.parser")


def extract_email(text: str) -> str:
    emails = re.findall(r"[\w.\-+]+@[\w.\-]+\.\w+", text)
    return unique_join(emails)


def extract_phone(text: str) -> str:
    phones = re.findall(r"(?:\+?86[- ]?)?(?:0\d{2,3}[- ]?)?\d{7,8}", text)
    return unique_join(phones[:3])


def clean_excel_cell(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_EXCEL_CHARS_RE.sub("", value)
    return value


def extract_urls(text: str) -> str:
    urls = re.findall(r"https?://[^\s，,；;。)）]+", text)
    return unique_join(urls)


def get_article_text(soup: BeautifulSoup) -> str:
    for selector in [
        ".wp_articlecontent",
        ".wp_article_content",
        ".article",
        ".entry",
        ".content",
        ".infobox",
        ".con2f",
        ".con1r",
        ".mainContent",
    ]:
        node = soup.select_one(selector)
        if node:
            return norm_text(node.get_text(" ", strip=True))
    return norm_text(soup.get_text(" ", strip=True))


def direction_from_text(text: str) -> str:
    text = norm_text(text)
    patterns = [
        r"(研究方向[为：:，, ]+[^。；;\n]{4,180})",
        r"(主要研究方向[为：:，, ]+[^。；;\n]{4,180})",
        r"(科研方向[为：:，, ]+[^。；;\n]{4,180})",
        r"(研究兴趣[为包括：:，, ]+[^。；;\n]{4,180})",
        r"(研究领域[为包括：:，, ]+[^。；;\n]{4,180})",
        r"((?:focus|focused|interests?|research interests?)[^。；;\n]{8,220})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            return norm_text(match.group(1))
    return ""


def title_from_name_or_category(name_text: str, category: str) -> str:
    text = norm_text(name_text)
    title_bits: list[str] = []
    inside = re.findall(r"[（(]([^）)]+)[）)]", text)
    for value in inside:
        title_bits.extend(re.split(r"[、,，/ ]+", value))
    if category and category not in {"跨学科博导"}:
        title_bits.insert(0, category)
    cleaned: list[str] = []
    for bit in title_bits:
        bit = norm_text(bit)
        if bit and bit not in {"博士"}:
            cleaned.append(bit)
    return unique_join(cleaned)


def clean_display_name(text: str) -> str:
    text = re.sub(r"[（(].*?[）)]", "", norm_text(text))
    text = re.sub(r"\s+", "", text)
    return clean_name(text)


def clean_faculty_name(text: Any) -> str:
    raw = norm_text(text)
    paren_chinese = re.search(r"[（(]([\u4e00-\u9fff·]{2,10})[）)]", raw)
    if paren_chinese:
        return clean_name(paren_chinese.group(1))
    leading_chinese = re.search(r"([\u4e00-\u9fff·]{2,10})", raw)
    if leading_chinese:
        return clean_name(leading_chinese.group(1))
    return raw


def infer_personal_homepage(soup: BeautifulSoup, profile_url: str) -> str:
    urls: list[str] = []
    profile_host = urlparse(profile_url).netloc
    for a in soup.select("a[href]"):
        href = urljoin(profile_url, a.get("href", ""))
        if not href.startswith(("http://", "https://")):
            continue
        host = urlparse(href).netloc
        text = norm_text(a.get_text(" ", strip=True)).lower()
        if host and host != profile_host:
            urls.append(href)
        elif any(token in text for token in ["主页", "homepage", "个人", "实验室"]):
            urls.append(href)
    return unique_join(urls)


def infer_clean_personal_homepage(soup: BeautifulSoup, profile_url: str) -> str:
    direct_urls: list[str] = []
    fallback_urls: list[str] = []
    profile = urlparse(profile_url)
    profile_host = profile.netloc.lower()
    profile_path = profile.path.rstrip("/").lower()
    for a in soup.select("a[href]"):
        raw_href = norm_text(a.get("href", "")).strip("，,；;。)）]")
        if not raw_href or raw_href in {"#", "--", "javascript:;"}:
            continue
        href = urljoin(profile_url, raw_href).strip("，,；;。)）]")
        if not href.startswith(("http://", "https://")):
            continue
        parsed = urlparse(href)
        host = parsed.netloc.lower()
        path = parsed.path.rstrip("/").lower()
        text = norm_text(a.get_text(" ", strip=True)).lower()
        if parsed.fragment.startswith(":~:text"):
            continue
        if host in GENERIC_COMMON_HOMEPAGE_HOSTS:
            continue
        if host in GENERIC_NON_HOMEPAGE_HOSTS or "dblp." in host:
            continue
        if host == profile_host and path == profile_path:
            continue
        anchor_is_homepage = any(
            token in text
            for token in ["个人主页", "个人网站", "主页", "homepage", "home page", "实验室", "课题组", "lab"]
        )
        obvious_personal_host = (
            host.endswith("github.io")
            or host in {"sites.google.com"}
            or "lab" in host
            or "group" in host
            or "faculty." in host
        )
        if anchor_is_homepage:
            direct_urls.append(href)
        elif obvious_personal_host:
            fallback_urls.append(href)
    return unique_join(direct_urls or fallback_urls)


def infer_personal_homepage_from_node(node: Any, profile_url: str) -> str:
    urls: list[str] = []
    if node is None:
        return ""
    profile_host = urlparse(profile_url).netloc
    common_hosts = {
        "www.nju.edu.cn",
        "ic.nju.edu.cn",
        "keysoftlab.nju.edu.cn",
        "cselab.nju.edu.cn",
        "software.nju.edu.cn",
        "cc.nju.edu.cn",
        "jw.nju.edu.cn",
    } | GENERIC_COMMON_HOMEPAGE_HOSTS
    for a in node.select("a[href]"):
        href = urljoin(profile_url, a.get("href", ""))
        if not href.startswith(("http://", "https://")):
            continue
        host = urlparse(href).netloc.lower()
        text = norm_text(a.get_text(" ", strip=True)).lower()
        if host in common_hosts or host in GENERIC_NON_HOMEPAGE_HOSTS or "dblp." in host:
            continue
        if host != profile_host or any(token in text for token in ["主页", "homepage", "个人", "实验室"]):
            urls.append(href)
    text_urls = extract_urls(node.get_text(" ", strip=True))
    if text_urls:
        urls.extend(
            url
            for url in text_urls.split("; ")
            if urlparse(url).netloc.lower() not in common_hosts
            and urlparse(url).netloc.lower() not in GENERIC_NON_HOMEPAGE_HOSTS
            and "dblp." not in urlparse(url).netloc.lower()
        )
    return unique_join(urls)


def score_first_pass_teacher(row: dict[str, Any]) -> dict[str, Any]:
    return score_teacher(row)


def fetch_sjtu_cs_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    payload = {
        "page": 1,
        "cat_id": "20",
        "cat_code": "jiaoshiml",
        "type": 1,
        "zm": "All",
        "zc": "全部",
        "search": "",
    }
    response = session.post(SJTU_CS_LIST_API_URL, data=payload, timeout=30)
    response.raise_for_status()
    soup = BeautifulSoup(response.json()["content"], "html.parser")

    merged: dict[str, dict[str, Any]] = {}
    order = 0

    def add_teacher(institute: str, name: str, href: str | None, directory_role: str) -> None:
        nonlocal order
        name = clean_name(name)
        if not name:
            return
        detail_url = urljoin(SJTU_CS_BASE_URL, href) if href else ""
        key = detail_url or f"NO_URL::{name}"
        if key not in merged:
            order += 1
            merged[key] = {
                "名录序号": order,
                "姓名": name,
                "教师主页链接": detail_url,
                "名录研究所": [],
                "所内职务": [],
            }
        merged[key]["名录研究所"].append(institute)
        if directory_role:
            merged[key]["所内职务"].append(f"{institute}:{directory_role}")

    for item in soup.select(".rc-item"):
        institute_node = item.select_one(".tit .name")
        institute = norm_text(institute_node.get_text(" ", strip=True) if institute_node else "")

        for p in item.select(".dt p"):
            p_text = norm_text(p.get_text(" ", strip=True))
            role = split_label_value(p_text)[0]
            for a in p.find_all("a"):
                add_teacher(institute, a.get_text(" ", strip=True), a.get("href"), role)

        for a in item.select(".name-list a"):
            add_teacher(institute, a.get_text(" ", strip=True), a.get("href"), "")

    directory_rows = list(merged.values())
    for row in directory_rows:
        row["名录研究所"] = unique_join(row["名录研究所"])
        row["所内职务"] = unique_join(row["所内职务"])
    directory_rows = sorted(directory_rows, key=lambda item: item["名录序号"])

    details: dict[str, dict[str, Any]] = {}
    urls = [row["教师主页链接"] for row in directory_rows if row.get("教师主页链接")]
    with ThreadPoolExecutor(max_workers=8) as executor:
        future_map = {executor.submit(parse_sjtu_cs_teacher_page, session, url): url for url in urls}
        for idx, future in enumerate(as_completed(future_map), start=1):
            url = future_map[future]
            details[url] = future.result()
            if idx % 40 == 0:
                time.sleep(0.5)

    rows: list[dict[str, Any]] = []
    for row in directory_rows:
        merged_row = dict(row)
        detail = details.get(row.get("教师主页链接", ""), {"抓取状态": "无详情页链接"})
        merged_row.update(detail)
        if merged_row.get("姓名_主页"):
            merged_row["姓名"] = clean_name(merged_row["姓名_主页"])
        merged_row.pop("姓名_主页", None)
        rows.append(merged_row)
    return rows


def parse_sjtu_cs_sections(soup: BeautifulSoup) -> dict[str, str]:
    sections: dict[str, str] = {}
    for item in soup.select(".js-dt .item"):
        title_node = item.select_one(".name")
        title = norm_text(title_node.get_text(" ", strip=True) if title_node else "")
        body_node = item.select_one(".txt")
        body = norm_text(body_node.get_text(" ", strip=True) if body_node else item.get_text(" ", strip=True))
        if title:
            sections[title] = body
    return sections


def extract_sjtu_cs_research_direction(sections: dict[str, str], full_text: str) -> str:
    def trim_candidate(text: str) -> str:
        text = norm_text(text)
        markers = [
            "教育背景",
            "工作履历",
            "教授课程",
            "论文发表",
            "项目",
            "奖励",
            "获奖",
            "学术服务",
            "邮箱：",
            "电话：",
            "地址：",
            "个人主页：",
            "所在研究所：",
            "Google Scholar",
            "ISE3302",
            "CS4303",
            "SE3303",
            "[Google",
        ]
        cut_points = [idx for marker in markers if (idx := text.find(marker, 20)) != -1]
        if cut_points:
            text = text[: min(cut_points)]
        return text[:700]

    direct_sections: list[str] = []
    for title, body in sections.items():
        if any(token in title for token in ["研究方向", "研究兴趣", "科研方向", "招生方向"]):
            direct_sections.append(trim_candidate(body))

    if direct_sections:
        return norm_text("；".join(direct_sections))[:1000]

    patterns = [
        "研究方向",
        "研究领域",
        "研究兴趣",
        "科研领域",
        "主要研究",
        "focus on",
        "focuses on",
        "specialize",
        "interests include",
    ]
    candidates: list[str] = []
    text_for_sentences = norm_text(sections.get("个人简介", "") or full_text)
    sentences = re.split(r"(?<=[。！？；;.!?])\s*", text_for_sentences)
    for sentence in sentences:
        lowered = sentence.lower()
        if any(pattern in lowered for pattern in patterns) or any(pattern in sentence for pattern in patterns):
            candidates.append(trim_candidate(sentence))
        if len(" ".join(candidates)) > 900:
            break

    if candidates:
        return norm_text(" ".join(candidates))[:1000]

    matched = [
        keyword
        for keyword, _ in KEYWORD_WEIGHTS
        if keyword_in_text(keyword, text_for_sentences.lower())
    ]
    if matched:
        return f"主页未单列研究方向；关键词匹配：{unique_join(matched[:12])}"

    return norm_text(sections.get("个人简介", ""))[:600]


def parse_sjtu_cs_teacher_page(session: requests.Session, url: str) -> dict[str, Any]:
    if not url:
        return {"抓取状态": "无详情页链接"}

    try:
        response = session.get(url, timeout=30)
        status = response.status_code
        if status != 200:
            return {"抓取状态": f"HTTP {status}"}
    except Exception as exc:  # noqa: BLE001
        return {"抓取状态": f"请求失败: {type(exc).__name__}"}

    soup = BeautifulSoup(response.text, "html.parser")
    info = soup.select_one(".js-info")
    sections = parse_sjtu_cs_sections(soup)

    parsed: dict[str, Any] = {"抓取状态": "OK"}
    if info:
        name_node = info.select_one(".name")
        title_node = info.select_one(".zw")
        parsed["姓名_主页"] = norm_text(name_node.get_text(" ", strip=True) if name_node else "")
        parsed["职称"] = norm_text(title_node.get_text(" ", strip=True) if title_node else "")

        details_texts = []
        homepages: list[str] = []
        for p in info.select(".dt p"):
            text = norm_text(p.get_text(" ", strip=True))
            details_texts.append(text)
            label, value = split_label_value(text)
            if label:
                if "邮箱" in label:
                    parsed["邮箱"] = value
                elif "电话" in label:
                    parsed["电话"] = value
                elif "地址" in label:
                    parsed["地址"] = value
                elif "所在研究所" in label:
                    parsed["主页研究所"] = value
                elif "个人主页" in label:
                    homepages.append(value)
            for a in p.find_all("a", href=True):
                homepages.append(a["href"])
        parsed["个人主页"] = unique_join(homepages)
        info_text = " ".join(details_texts)
    else:
        title_text = soup.title.get_text(" ", strip=True) if soup.title else ""
        parsed["姓名_主页"] = title_text.split("-", 1)[0].strip()
        info_text = ""

    full_text = norm_text(" ".join(sections.values()))
    all_text = norm_text(f"{info_text} {full_text}")
    if "邮箱" not in parsed:
        email_match = re.search(r"[\w.\-+]+@[\w.\-]+\.\w+", all_text)
        parsed["邮箱"] = email_match.group(0) if email_match else ""

    parsed["研究方向"] = extract_sjtu_cs_research_direction(sections, all_text)
    parsed["个人简介摘要"] = norm_text(sections.get("个人简介", ""))[:1000]
    parsed["详情页全文片段"] = all_text[:2500]
    return parsed


def parse_sjtu_ai_detail(session: requests.Session, row: dict[str, Any]) -> dict[str, Any]:
    soup = fetch(session, row["教师主页链接"], "utf-8")
    text = norm_text(soup.get_text(" ", strip=True))
    detail = dict(row)
    detail["抓取状态"] = "成功"

    name = row["姓名"]
    if not name:
        # The detail page usually places the name before "职称：".
        match = re.search(r"专职教师\s+([^\s]{2,8})\s+职称[:：]", text)
        if match:
            detail["姓名"] = clean_name(match.group(1))

    title_match = re.search(r"职称[:：]\s*([^邮箱个人主页]{2,40})", text)
    if title_match:
        detail["职称"] = norm_text(title_match.group(1))

    detail["邮箱"] = extract_email(text)
    phone = extract_phone(text)
    if phone and "021" not in phone:
        detail["电话"] = phone

    homepage = row.get("个人主页", "")
    if not homepage:
        hp_match = re.search(r"个人主页[:：]\s*(https?://\S+)", text)
        homepage = hp_match.group(1) if hp_match else ""
    detail["个人主页"] = norm_text(homepage)

    summary = ""
    if "个人简介" in text:
        summary = text.split("个人简介", 1)[1]
        summary = re.split(r"徐汇区办公地址|闵行区办公地址|版权所有", summary, 1)[0]
    detail["个人简介摘要"] = norm_text(summary or text)[:1200]
    detail["研究方向"] = direction_from_text(detail["个人简介摘要"])
    detail["主页研究所"] = "人工智能学院"
    return detail


def fetch_sjtu_ai_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    merged: dict[str, dict[str, Any]] = {}

    for a in soup.select('a[href*="/cn/facultydetails/zzjs/"]'):
        href = urljoin(config.directory_url, a.get("href", ""))
        text = clean_display_name(a.get_text(" ", strip=True))
        if not text:
            continue
        if href not in merged:
            merged[href] = {
                "姓名": text,
                "教师主页链接": href,
                "名录研究所": "上海交通大学人工智能学院",
                "所内职务": "专职教师",
                "职称": "",
                "邮箱": "",
                "电话": "",
                "地址": "",
                "个人主页": "",
                "主页研究所": "人工智能学院",
                "研究方向": "",
                "个人简介摘要": "",
                "抓取状态": "待抓取",
            }

    rows: list[dict[str, Any]] = []
    for index, row in enumerate(merged.values(), start=1):
        row["名录序号"] = index
        try:
            rows.append(parse_sjtu_ai_detail(session, row))
        except Exception as exc:
            row["抓取状态"] = f"失败：{exc}"
            rows.append(row)
        time.sleep(0.05)
    return rows


def parse_nju_detail(session: requests.Session, row: dict[str, Any]) -> dict[str, Any]:
    detail = dict(row)
    profile_url = row["教师主页链接"]
    profile_host = urlparse(profile_url).netloc.lower()
    official_hosts = {"cs.nju.edu.cn", "ai.nju.edu.cn"}
    if profile_host not in official_hosts:
        try:
            soup = fetch(session, profile_url, "utf-8")
            text = get_article_text(soup)
            if len(text) < 50:
                text = norm_text(soup.get_text(" ", strip=True))
            detail["抓取状态"] = "外部主页-已解析"
            detail["邮箱"] = extract_email(text)
            detail["个人主页"] = profile_url
            detail["个人简介摘要"] = text[:1200]
            detail["研究方向"] = direction_from_text(text)
        except Exception as exc:
            detail["抓取状态"] = f"外部主页-解析失败：{exc}"
            detail["个人主页"] = profile_url
            detail["个人简介摘要"] = ""
        detail["个人主页"] = profile_url
        return detail

    soup = fetch(session, profile_url, "utf-8")
    article_node = None
    for selector in [".wp_articlecontent", ".wp_article_content", ".article", ".entry", ".content", ".infobox"]:
        article_node = soup.select_one(selector)
        if article_node:
            break
    article_text = get_article_text(soup)
    full_text = norm_text(soup.get_text(" ", strip=True))
    detail["抓取状态"] = "成功"
    detail["邮箱"] = extract_email(full_text)
    detail["电话"] = extract_phone(full_text)
    detail["个人主页"] = infer_personal_homepage_from_node(article_node, profile_url) or infer_personal_homepage(soup, profile_url)
    detail["个人简介摘要"] = article_text[:1200]
    detail["研究方向"] = direction_from_text(article_text)
    detail["主页研究所"] = row.get("主页研究所") or row.get("名录研究所") or "南京大学"
    return detail


def fetch_nju_category(session: requests.Session, config: TargetConfig, category: str, url: str) -> list[dict[str, Any]]:
    soup = fetch(session, url, "utf-8")
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    selectors = [".wp_article_list a[href]", ".list a[href]", "a[href]"]
    anchors = []
    for selector in selectors:
        anchors = soup.select(selector)
        if anchors:
            break
    for a in anchors:
        text = norm_text(a.get_text(" ", strip=True))
        href = urljoin(url, a.get("href", ""))
        if not text or text in {category, "首页"}:
            continue
        if not ("/page.htm" in href or "_redirect" in href or href.startswith("https://www.nju.edu.cn") or href.startswith("https://sme.nju.edu.cn") or "github.io" in href):
            continue
        name = clean_display_name(text)
        if not name or len(name) < 2:
            continue
        key = f"{name}::{href}"
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "姓名": name,
                "职称": title_from_name_or_category(text, category),
                "教师主页链接": href,
                "名录研究所": "南京大学计算机学院",
                "所内职务": category,
                "邮箱": "",
                "电话": "",
                "地址": "",
                "个人主页": "",
                "主页研究所": "",
                "研究方向": "",
                "个人简介摘要": "",
                "抓取状态": "待抓取",
            }
        )
    return rows


def fetch_nju_cs_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for category, url in NJU_ACADEMIC_CATEGORIES:
        for row in fetch_nju_category(session, config, category, url):
            key = row["教师主页链接"] or row["姓名"]
            if key in merged:
                merged[key]["所内职务"] = unique_join([merged[key].get("所内职务", ""), row.get("所内职务", "")])
                if row.get("职称") and row["职称"] not in merged[key].get("职称", ""):
                    merged[key]["职称"] = unique_join([merged[key].get("职称", ""), row["职称"]])
            else:
                merged[key] = row

    rows: list[dict[str, Any]] = []
    for index, row in enumerate(merged.values(), start=1):
        row["名录序号"] = index
        try:
            rows.append(parse_nju_detail(session, row))
        except Exception as exc:
            row["抓取状态"] = f"失败：{exc}"
            rows.append(row)
        time.sleep(0.05)
    return rows


def clean_nju_ai_display_name(text: str) -> str:
    raw = re.sub(r"[（(].*?[）)]", "", norm_text(text)).strip()
    if re.search(r"[\u4e00-\u9fff]", raw):
        return clean_faculty_name(raw)
    return re.sub(r"\s+", " ", raw)


def fetch_nju_ai_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()

    for news_list in soup.select("ul.news_list"):
        heading = news_list.find_previous(class_="subcolumn-name")
        section = norm_text(heading.get_text(" ", strip=True) if heading else "")
        if section not in NJU_AI_ACADEMIC_SECTIONS:
            continue
        for a in news_list.select("a[href]"):
            text = norm_text(a.get_text(" ", strip=True) or a.get("title", ""))
            href = urljoin(config.directory_url, a.get("href", ""))
            name = clean_nju_ai_display_name(text)
            if not name or len(name) < 2:
                continue
            key = f"{name}::{href}"
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "姓名": name,
                    "职称": title_from_name_or_category(text, section),
                    "教师主页链接": href,
                    "名录研究所": "南京大学人工智能学院",
                    "所内职务": section,
                    "邮箱": "",
                    "电话": "",
                    "地址": "",
                    "个人主页": "" if urlparse(href).netloc.endswith("nju.edu.cn") else href,
                    "主页研究所": "南京大学人工智能学院",
                    "研究方向": "",
                    "个人简介摘要": "",
                    "抓取状态": "待抓取",
                }
            )

    output: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        row["名录序号"] = index
        try:
            output.append(parse_nju_detail(session, row))
        except Exception as exc:
            row["抓取状态"] = f"失败：{exc}"
            output.append(row)
        time.sleep(0.05)
    return output


def section_text_without_heading(section: Any, heading: Any) -> str:
    section_text = norm_text(section.get_text(" ", strip=True))
    heading_text = norm_text(heading.get_text(" ", strip=True) if heading else "")
    if heading_text and section_text.startswith(heading_text):
        section_text = section_text[len(heading_text) :]
    return norm_text(section_text)


def extract_ruc_sections(soup: BeautifulSoup) -> dict[str, str]:
    sections: dict[str, str] = {}
    for section in soup.select("section"):
        heading = section.select_one("h2.colorlib-heading, h2, h3")
        if not heading:
            continue
        title = norm_text(heading.get_text(" ", strip=True))
        body = section_text_without_heading(section, heading)
        if title and body:
            sections[title] = body
    return sections


def infer_ruc_personal_homepage(soup: BeautifulSoup, profile_url: str) -> str:
    direct_urls: list[str] = []
    fallback_urls: list[str] = []
    profile_host = urlparse(profile_url).netloc.lower()
    for a in soup.select("a[href]"):
        raw_href = norm_text(a.get("href", "")).strip("，,；;。)）]")
        if not raw_href or raw_href in {"#", "--", "javascript:;"}:
            continue
        href = urljoin(profile_url, raw_href)
        href = href.strip("，,；;。)）]")
        if not href.startswith(("http://", "https://")):
            continue
        parsed = urlparse(href)
        host = parsed.netloc.lower()
        text = norm_text(a.get_text(" ", strip=True)).lower()
        path = parsed.path.lower()
        if parsed.fragment.startswith(":~:text"):
            continue
        if host == profile_host and ("/english/" in path or path == urlparse(profile_url).path.lower()):
            continue
        in_profile_sidebar = a.find_parent("aside") is not None
        homepage_anchor = (
            any(token in text for token in ["个人主页", "点击访问"])
            or ("homepage" in text and in_profile_sidebar)
            or ("实验室" in text and in_profile_sidebar)
            or ("lab" in text and in_profile_sidebar)
        )
        if host in RUC_COMMON_HOSTS and not homepage_anchor:
            continue
        if "教师入口" in text or "teacher entry" in text:
            continue
        if host in RUC_NON_HOMEPAGE_HOSTS or "dblp." in host:
            continue
        if homepage_anchor:
            direct_urls.append(href)
            continue
        if host.endswith("github.io") or host in {"sites.google.com"} or "lab" in host or "lab" in path:
            fallback_urls.append(href)

    profile_text_parts = [
        norm_text(node.get_text(" ", strip=True))
        for node in soup.select(".text-desc")
    ]
    for section in soup.select("section"):
        heading = section.select_one("h2.colorlib-heading, h2, h3")
        heading_text = norm_text(heading.get_text(" ", strip=True) if heading else "")
        if any(token in heading_text for token in ["研究方向", "研究兴趣", "科研方向", "研究领域"]):
            profile_text_parts.append(section_text_without_heading(section, heading))
    for text_part in profile_text_parts:
        for match in re.finditer(r"(?:个人主页|详见个人主页)[：:\s]+(https?://[^\s，,；;。)）]+)", text_part):
            href = match.group(1).strip("，,；;。)）]")
            host = urlparse(href).netloc.lower()
            if host not in RUC_NON_HOMEPAGE_HOSTS and "dblp." not in host:
                direct_urls.append(href)

    return unique_join(direct_urls or fallback_urls)


def extract_ruc_research_direction(sections: dict[str, str], summary: str, full_text: str) -> str:
    for title, body in sections.items():
        if any(token in title for token in ["研究方向", "研究兴趣", "科研方向", "研究领域"]):
            return norm_text(body)[:1000]
    candidate = direction_from_text(summary) or direction_from_text(full_text)
    if candidate:
        return candidate[:1000]
    return summary[:600]


def parse_ruc_gsai_detail(session: requests.Session, row: dict[str, Any]) -> dict[str, Any]:
    detail = dict(row)
    profile_url = row["教师主页链接"]
    soup = fetch(session, profile_url, "utf-8")
    full_text = norm_text(soup.get_text(" ", strip=True))
    sections = extract_ruc_sections(soup)

    name_node = soup.select_one("aside h1, h1")
    if name_node:
        detail["姓名"] = clean_name(name_node.get_text(" ", strip=True))

    positions = [
        norm_text(node.get_text(" ", strip=True))
        for node in soup.select("aside .position, .position")
        if norm_text(node.get_text(" ", strip=True))
    ]
    if positions:
        detail["职称"] = unique_join(positions[-2:]) if len(positions) > 1 else positions[0]

    summary_node = soup.select_one(".text-desc")
    summary = norm_text(summary_node.get_text(" ", strip=True) if summary_node else "")
    detail["邮箱"] = extract_email(full_text)
    detail["电话"] = extract_phone(full_text)
    detail["个人主页"] = infer_ruc_personal_homepage(soup, profile_url)
    detail["主页研究所"] = "高瓴人工智能学院"
    detail["研究方向"] = extract_ruc_research_direction(sections, summary, full_text)

    summary_parts = [summary, detail["研究方向"]]
    if sections.get("研究成果"):
        summary_parts.append(sections["研究成果"][:700])
    detail["个人简介摘要"] = unique_join(summary_parts)[:1200]
    detail["抓取状态"] = "成功"
    return detail


def fetch_ruc_gsai_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()

    for h2 in soup.find_all("h2"):
        a = h2.find("a", href=True)
        if not a:
            continue
        href = urljoin(config.directory_url, a["href"])
        if "gsai.ruc.edu.cn" not in urlparse(href).netloc:
            continue
        name = clean_display_name(a.get_text(" ", strip=True))
        if name in RUC_GSAI_NAME_OVERRIDES:
            name = RUC_GSAI_NAME_OVERRIDES[name]
        if not name:
            continue
        title_node = h2.find("small")
        title = norm_text(title_node.get_text(" ", strip=True) if title_node else "")
        summary_node = h2.find_next("p", class_="position") or h2.find_next("p")
        summary = norm_text(summary_node.get_text(" ", strip=True) if summary_node else "")
        key = f"{name}::{href}"
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "名录序号": len(rows) + 1,
                "姓名": name,
                "职称": title,
                "教师主页链接": href,
                "名录研究所": "中国人民大学高瓴人工智能学院",
                "所内职务": "教师队伍",
                "邮箱": "",
                "电话": "",
                "地址": "",
                "个人主页": "",
                "主页研究所": "高瓴人工智能学院",
                "研究方向": direction_from_text(summary),
                "个人简介摘要": summary[:1200],
                "抓取状态": "待抓取",
            }
        )

    parsed_rows: list[dict[str, Any]] = []
    for row in rows:
        try:
            parsed_rows.append(parse_ruc_gsai_detail(session, row))
        except Exception as exc:
            row["抓取状态"] = f"失败：{exc}"
            parsed_rows.append(row)
        time.sleep(0.05)
    return parsed_rows


def valid_http_url(url: str) -> bool:
    return norm_text(url).startswith(("http://", "https://")) and norm_text(url) != "#"


def first_http_url(value: Any) -> str:
    match = re.search(r"https?://[^\s，,；;。)）]+", norm_text(value))
    return match.group(0).strip("，,；;。)）]") if match else ""


def strip_at_markers(text: str, markers: list[str]) -> str:
    result = norm_text(text)
    for marker in markers:
        index = result.find(marker)
        if index > 0:
            result = result[:index]
    return norm_text(result)


def extract_labeled_value(text: str, label: str, stops: list[str]) -> str:
    source = norm_text(text)
    index = source.find(label)
    if index < 0:
        return ""
    start = index + len(label)
    while start < len(source) and source[start] in " ：:\t":
        start += 1
    end = len(source)
    for stop in stops:
        stop_index = source.find(stop, start)
        if stop_index >= start:
            end = min(end, stop_index)
    return norm_text(source[start:end]).strip(" ：:")


def post_general_query(
    session: requests.Session,
    url: str,
    data: dict[str, Any],
) -> dict[str, Any]:
    response = session.post(url, data=data, timeout=30, headers=HEADERS)
    response.raise_for_status()
    return response.json()


def parse_generic_profile_detail(
    session: requests.Session,
    row: dict[str, Any],
    homepage_as_personal: bool = False,
) -> dict[str, Any]:
    detail = dict(row)
    profile_url = norm_text(row.get("教师主页链接"))
    if not valid_http_url(profile_url):
        detail["抓取状态"] = "无独立主页"
        detail["个人简介摘要"] = norm_text(row.get("个人简介摘要"))[:1200]
        detail["研究方向"] = norm_text(row.get("研究方向")) or direction_from_text(detail["个人简介摘要"])
        return detail

    try:
        soup = fetch(session, profile_url, "utf-8")
    except Exception as exc:  # noqa: BLE001
        detail["抓取状态"] = f"主页抓取失败：{exc}"
        return detail

    article_text = get_article_text(soup)
    full_text = norm_text(soup.get_text(" ", strip=True))
    if len(article_text) < 50:
        article_text = full_text
    detail["抓取状态"] = "成功"
    detail["邮箱"] = norm_text(row.get("邮箱")) or extract_email(full_text)
    detail["电话"] = norm_text(row.get("电话")) or extract_phone(full_text)
    homepage = norm_text(row.get("个人主页"))
    if not homepage:
        homepage = infer_clean_personal_homepage(soup, profile_url)
    if homepage_as_personal and not homepage:
        homepage = profile_url
    detail["个人主页"] = homepage
    detail["个人简介摘要"] = unique_join([row.get("个人简介摘要", ""), article_text])[:1200]
    detail["研究方向"] = norm_text(row.get("研究方向")) or direction_from_text(article_text) or direction_from_text(full_text)
    return detail


RUC_INFO_DEPARTMENTS = {
    "dsjkxygcjys": "大数据科学与工程教研室",
    "jjxxglx1": "经济信息管理系",
    "jsjkxyjsx1": "计算机科学与技术系",
    "xxjsjcjys1": "信息技术基础教研室",
}


def ruc_info_department_from_url(url: str) -> str:
    path = urlparse(url).path
    if "/rtjs/" in path:
        return "荣退教师"
    for slug, department in RUC_INFO_DEPARTMENTS.items():
        if f"/{slug}/" in path:
            return department
    return "中国人民大学信息学院"


def ruc_info_page_index(url: str) -> int:
    match = re.search(r"/index(\d*)\.htm$", urlparse(url).path)
    if not match or not match.group(1):
        return 0
    return int(match.group(1))


def discover_ruc_info_pages(session: requests.Session, config: TargetConfig) -> list[str]:
    soup = fetch(session, config.directory_url, "utf-8")
    pages = {config.directory_url}
    for a in soup.select("a[href]"):
        href = urljoin(config.directory_url, a.get("href", ""))
        parsed = urlparse(href)
        if parsed.netloc != "info.ruc.edu.cn":
            continue
        if re.search(r"/jsky/szdw/ajxjgcx/bx/bx1/index\d*\.htm$", parsed.path):
            pages.add(href)
    return sorted(pages, key=ruc_info_page_index)


def parse_ruc_info_card(card: Any, page_url: str) -> dict[str, Any] | None:
    a = card.select_one("a[href]")
    if not a:
        return None
    href = urljoin(page_url, a.get("href", ""))
    name = clean_name(card.select_one(".text1").get_text(" ", strip=True) if card.select_one(".text1") else "")
    if not name:
        return None

    labels = [norm_text(node.get_text(" ", strip=True)) for node in card.select(".research-text .text2")]
    values = [norm_text(node.get_text(" ", strip=True)) for node in card.select(".research-text .text3")]
    card_info = dict(zip(labels, values, strict=False))
    direction = norm_text(card_info.get("研究方向"))
    courses = norm_text(card_info.get("讲授课程"))
    department = ruc_info_department_from_url(href)
    summary = unique_join(
        [
            f"研究方向：{direction}" if direction else "",
            f"讲授课程：{courses}" if courses else "",
        ]
    )
    return {
        "姓名": name,
        "职称": "",
        "教师主页链接": href,
        "名录研究所": department,
        "官方系别": department,
        "所内职务": "师资队伍",
        "邮箱": "",
        "电话": "",
        "地址": "",
        "个人主页": "",
        "主页研究所": department,
        "研究方向": direction,
        "个人简介摘要": summary[:1200],
        "抓取状态": "待抓取",
    }


def extract_ruc_info_article_text(soup: BeautifulSoup, name: str) -> str:
    article_text = get_article_text(soup)
    for marker in ["详细资料 Details", "Details"]:
        marker_index = article_text.find(marker)
        if marker_index >= 0:
            article_text = article_text[marker_index + len(marker) :]
            break
    article_text = norm_text(article_text)
    if name and article_text.startswith(name):
        article_text = norm_text(article_text[len(name) :])
    return article_text


def extract_ruc_info_personal_homepage(article_text: str, soup: BeautifulSoup, profile_url: str) -> str:
    labeled = extract_labeled_value(
        article_text,
        "个人主页",
        ["更多", "教育经历", "工作经历", "研究方向", "指导学生", "讲授课程", "科研项目", "论文", "奖励"],
    )
    homepage = first_http_url(labeled)
    host = urlparse(homepage).netloc.lower() if homepage else ""
    if homepage and host not in GENERIC_NON_HOMEPAGE_HOSTS and host not in RUC_NON_HOMEPAGE_HOSTS and "dblp." not in host:
        return homepage
    return infer_clean_personal_homepage(soup, profile_url)


def parse_ruc_info_detail(session: requests.Session, row: dict[str, Any]) -> dict[str, Any]:
    detail = dict(row)
    profile_url = norm_text(row.get("教师主页链接"))
    soup = fetch(session, profile_url, "utf-8")
    title_text = norm_text(soup.title.get_text(" ", strip=True) if soup.title else "")
    title_parts = [norm_text(part) for part in re.split(r"\s*-\s*", title_text) if norm_text(part)]

    if title_parts and title_parts[0] != "不限":
        detail["姓名"] = clean_name(title_parts[0])
    if len(title_parts) > 1 and title_parts[1] != "不限":
        detail["职称"] = title_parts[1]
    if len(title_parts) > 2 and title_parts[2] != "不限":
        detail["主页研究所"] = title_parts[2]
        detail["官方系别"] = title_parts[2]

    name = norm_text(detail.get("姓名"))
    article_text = extract_ruc_info_article_text(soup, name)
    full_text = norm_text(soup.get_text(" ", strip=True))
    detail["邮箱"] = norm_text(row.get("邮箱")) or extract_email(article_text or full_text)
    detail["电话"] = norm_text(row.get("电话")) or extract_phone(article_text or full_text)
    detail["个人主页"] = norm_text(row.get("个人主页")) or extract_ruc_info_personal_homepage(article_text, soup, profile_url)
    detail["研究方向"] = norm_text(row.get("研究方向")) or direction_from_text(article_text) or direction_from_text(full_text)
    detail["个人简介摘要"] = unique_join([row.get("个人简介摘要", ""), article_text])[:1200]
    detail["抓取状态"] = "成功"
    return detail


def fetch_ruc_info_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page_url in discover_ruc_info_pages(session, config):
        soup = fetch(session, page_url, "utf-8")
        for card in soup.select(".research"):
            row = parse_ruc_info_card(card, page_url)
            if not row:
                continue
            key = f"{row['姓名']}::{row['教师主页链接']}"
            if key in seen:
                continue
            seen.add(key)
            row["名录序号"] = len(rows) + 1
            rows.append(row)

    parsed_rows: list[dict[str, Any]] = []
    for row in rows:
        try:
            parsed_rows.append(parse_ruc_info_detail(session, row))
        except Exception as exc:  # noqa: BLE001
            row["抓取状态"] = f"失败：{exc}"
            parsed_rows.append(row)
        time.sleep(0.05)
    return parsed_rows


def fetch_ruc_ssai_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    rows: list[dict[str, Any]] = []
    for card in soup.select(".product-card"):
        text = norm_text(card.get_text(" ", strip=True))
        if not text:
            continue
        name_match = re.match(r"([\u4e00-\u9fff·]{2,8})\s+", text)
        if not name_match:
            continue
        name = clean_name(name_match.group(1))
        remainder = norm_text(text[name_match.end() :])
        direction_match = re.search(r"研究方向[：:]\s*(.*?)(?:邮箱[：:]|$)", remainder)
        email = extract_email(text)
        direction = norm_text(direction_match.group(1)) if direction_match else direction_from_text(text)
        title = remainder
        title = re.split(r"研究方向[：:]|邮箱[：:]", title, 1)[0]
        title = norm_text(title)
        rows.append(
            {
                "名录序号": len(rows) + 1,
                "姓名": name,
                "职称": title,
                "教师主页链接": f"{config.directory_url}#{name}",
                "名录研究所": "中国人民大学苏州人工智能学院",
                "所内职务": "师资队伍",
                "邮箱": email,
                "电话": "",
                "地址": "",
                "个人主页": config.directory_url,
                "主页研究所": "苏州人工智能学院",
                "研究方向": direction,
                "个人简介摘要": text[:1200],
                "抓取状态": "成功-目录卡片",
            }
        )
    return rows


def parse_nju_ra_anchor(a: Any, config: TargetConfig) -> dict[str, Any] | None:
    href = urljoin(config.directory_url, a.get("href", ""))
    if "/szll/zzjs/" not in href or not href.endswith(".html"):
        return None
    text = norm_text(a.get_text(" ", strip=True))
    if "职称" not in text and "研究方向" not in text:
        return None
    name_match = re.match(r"([\u4e00-\u9fff·]{2,8})\s*", text)
    if not name_match:
        return None
    name = clean_name(name_match.group(1))
    title = ""
    title_match = re.search(r"职称[：:]\s*(.*?)(?:研究方向[：:]|$)", text)
    if title_match:
        title = norm_text(title_match.group(1))
    position = norm_text(text[name_match.end() : text.find("职称") if "职称" in text else len(text)])
    direction = ""
    direction_match = re.search(r"研究方向[：:]\s*(.+)$", text)
    if direction_match:
        direction = norm_text(direction_match.group(1))
    return {
        "姓名": name,
        "职称": title,
        "教师主页链接": href,
        "名录研究所": "南京大学机器人与自动化学院",
        "所内职务": position or "专职教师",
        "邮箱": "",
        "电话": "",
        "地址": "",
        "个人主页": "",
        "主页研究所": "机器人与自动化学院",
        "研究方向": direction,
        "个人简介摘要": text[:1200],
        "抓取状态": "待抓取",
    }


def fetch_nju_ra_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    merged: dict[str, dict[str, Any]] = {}
    for a in soup.select('a[href*="/szll/zzjs/"]'):
        row = parse_nju_ra_anchor(a, config)
        if not row:
            continue
        merged[row["教师主页链接"]] = row

    rows: list[dict[str, Any]] = []
    for index, row in enumerate(merged.values(), start=1):
        row["名录序号"] = index
        detail = parse_generic_profile_detail(session, row)
        if not detail.get("研究方向"):
            detail["研究方向"] = row.get("研究方向", "")
        rows.append(detail)
        time.sleep(0.05)
    return rows


def fetch_nju_is_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    api_url = "https://is.nju.edu.cn/_wp3services/generalQuery?queryObj=teacherHome"
    return_infos = [
        {"field": field, "name": field}
        for field in ["headerPic", "exField1", "exField2", "exField3", "Phone", "cnUrl", "title", "post", "phone", "siteSort"]
    ]
    rows_by_key: dict[str, dict[str, Any]] = {}
    for page_index in range(1, 20):
        data = {
            "siteId": 786,
            "pageIndex": page_index,
            "rows": 12,
            "orders": json.dumps([{"field": "siteSort", "type": "asc"}], ensure_ascii=False),
            "returnInfos": json.dumps(return_infos, ensure_ascii=False),
            "conditions": json.dumps([{"field": "published", "value": "1", "judge": "="}], ensure_ascii=False),
            "articleType": 1,
            "level": 1,
        }
        payload = post_general_query(session, api_url, data)
        page_rows = payload.get("data", [])
        if not page_rows and page_index > 1:
            break
        for item in page_rows:
            title = norm_text(item.get("title"))
            academic_title = norm_text(item.get("exField2"))
            if not title or academic_title not in NJU_IS_ACADEMIC_TITLES:
                continue
            profile_url = first_http_url(item.get("cnUrl"))
            row = {
                "姓名": clean_name(title),
                "职称": academic_title,
                "教师主页链接": profile_url,
                "名录研究所": "南京大学智能科学与技术学院",
                "所内职务": academic_title,
                "邮箱": extract_email(" ".join(norm_text(item.get(key)) for key in ["Phone", "phone", "exField3"])),
                "电话": extract_phone(" ".join(norm_text(item.get(key)) for key in ["Phone", "phone", "exField3"])),
                "地址": "",
                "个人主页": "",
                "主页研究所": "智能科学与技术学院",
                "研究方向": norm_text(item.get("exField1")),
                "个人简介摘要": unique_join([item.get("exField1", ""), item.get("exField3", "")])[:1200],
                "抓取状态": "待抓取",
            }
            rows_by_key[profile_url or title] = row
        time.sleep(0.05)

    rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows_by_key.values(), start=1):
        row["名录序号"] = index
        rows.append(parse_generic_profile_detail(session, row))
        time.sleep(0.05)
    return rows


def parse_nju_ic_detail(session: requests.Session, row: dict[str, Any]) -> dict[str, Any]:
    detail = dict(row)
    profile_url = row["教师主页链接"]
    soup = fetch(session, profile_url, "utf-8")
    article_node = soup.select_one(".article") or soup.select_one(".wp_articlecontent") or soup.select_one(".wp_article_content")
    article_text = norm_text(article_node.get_text(" ", strip=True) if article_node else soup.get_text(" ", strip=True))
    detail["抓取状态"] = "成功"

    name = extract_labeled_value(article_text, "姓名", ["职称", "联系方式", "电子邮箱", "个人简介", "研究方向"])
    if name:
        detail["姓名"] = clean_name(name)

    title = extract_labeled_value(article_text, "职称", ["联系方式", "电子邮箱", "个人简介", "研究方向"])
    if title:
        detail["职称"] = title

    email_text = extract_labeled_value(article_text, "电子邮箱", ["个人简介", "研究方向"])
    detail["邮箱"] = extract_email(email_text) or extract_email(article_text)
    detail["电话"] = extract_phone(article_text)
    detail["个人主页"] = infer_personal_homepage_from_node(article_node, profile_url)

    summary = extract_labeled_value(article_text, "个人简介", ["研究方向"]) or article_text
    direction = extract_labeled_value(article_text, "研究方向", []) or direction_from_text(article_text)
    detail["个人简介摘要"] = summary[:1200]
    detail["研究方向"] = direction
    detail["主页研究所"] = "集成电路学院"
    return detail


def fetch_nju_ic_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    pending = [urljoin(config.directory_url, "/qyml/list.htm")]
    visited: set[str] = set()
    merged: dict[str, dict[str, Any]] = {}

    while pending:
        page_url = pending.pop(0)
        if page_url in visited:
            continue
        visited.add(page_url)
        soup = fetch(session, page_url, "utf-8")
        for li in soup.select(".news_list li"):
            a = li.select_one('a[href*="/page.htm"]')
            if not a:
                continue
            href = urljoin(page_url, a.get("href", ""))
            text = norm_text(li.get_text(" ", strip=True))
            name = clean_faculty_name(text)
            if not name:
                continue
            title = norm_text(text.replace(name, "", 1))
            merged[href] = {
                "姓名": name,
                "职称": title,
                "教师主页链接": href,
                "名录研究所": "南京大学集成电路学院",
                "所内职务": title or "全院名录",
                "邮箱": "",
                "电话": "",
                "地址": "",
                "个人主页": "",
                "主页研究所": "集成电路学院",
                "研究方向": "",
                "个人简介摘要": "",
                "抓取状态": "待抓取",
            }

        for a in soup.select(".wp_paging a[href]"):
            text = norm_text(a.get_text(" ", strip=True))
            href = urljoin(page_url, a.get("href", ""))
            if "下一页" in text and href.startswith(("http://", "https://")) and href not in visited:
                pending.append(href)

    rows: list[dict[str, Any]] = []
    for index, row in enumerate(merged.values(), start=1):
        row["名录序号"] = index
        try:
            rows.append(parse_nju_ic_detail(session, row))
        except Exception as exc:  # noqa: BLE001
            row["抓取状态"] = f"失败：{exc}"
            rows.append(row)
        time.sleep(0.05)
    return rows


def fetch_fudan_ciram_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    api_url = "https://ciram.fudan.edu.cn/_wp3services/generalQuery?queryObj=articles"
    fields = ["title", "f9", "f1", "f2", "f4", "shortTitle", "subTitle", "imgPath", "letter", "url"]
    data = {
        "siteId": 1083,
        "columnId": 50802,
        "pageIndex": 1,
        "rows": 300,
        "orders": json.dumps([{"field": "letter", "type": "asc"}], ensure_ascii=False),
        "returnInfos": json.dumps([{"field": field, "name": field} for field in fields], ensure_ascii=False),
        "conditions": json.dumps([{"field": "scope", "value": 0, "judge": "="}], ensure_ascii=False),
    }
    payload = post_general_query(session, api_url, data)
    rows: list[dict[str, Any]] = []
    for item in payload.get("data", []):
        short_title = norm_text(item.get("shortTitle"))
        unit = norm_text(item.get("f4"))
        if short_title == "党政管理" or unit in FUDAN_CIRAM_ADMIN_UNITS:
            continue
        name = clean_faculty_name(item.get("title"))
        if not name:
            continue
        raw_url = first_http_url(item.get("url"))
        profile_url = raw_url if valid_http_url(raw_url) else f"{config.directory_url}#{name}"
        raw_host = urlparse(raw_url).netloc.lower() if raw_url else ""
        personal_homepage = raw_url if valid_http_url(raw_url) and raw_host not in GENERIC_NON_HOMEPAGE_HOSTS else ""
        title = unique_join([item.get("f9", ""), item.get("f1", "")])
        direction = norm_text(item.get("f2"))
        rows.append(
            {
                "名录序号": len(rows) + 1,
                "姓名": name,
                "职称": title,
                "教师主页链接": profile_url,
                "名录研究所": "复旦大学智能机器人与先进制造创新学院",
                "所内职务": unique_join([short_title or "教学科研", unit]),
                "邮箱": "",
                "电话": "",
                "地址": "",
                "个人主页": personal_homepage,
                "主页研究所": unit or "智能机器人与先进制造创新学院",
                "研究方向": direction,
                "个人简介摘要": unique_join([item.get("f1", ""), item.get("f2", ""), item.get("f4", ""), item.get("subTitle", "")])[:1200],
                "抓取状态": "待抓取" if valid_http_url(profile_url) and "#" not in profile_url else "成功-接口条目",
            }
        )

    parsed_rows: list[dict[str, Any]] = []
    for row in rows:
        if valid_http_url(row["教师主页链接"]) and "#" not in row["教师主页链接"]:
            parsed_rows.append(parse_generic_profile_detail(session, row))
            time.sleep(0.05)
        else:
            parsed_rows.append(row)
    return parsed_rows


def fetch_fudan_ai_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    api_url = "https://ai.fudan.edu.cn/_wp3services/generalQuery?queryObj=teacherHome"
    return_infos = [
        {"field": field, "name": field}
        for field in ["title", "cnUrl", "headerPic", "exField1", "exField7", "exField9", "email", "career"]
    ]
    data = {
        "siteId": 577,
        "pageIndex": 1,
        "rows": 999,
        "conditions": json.dumps([], ensure_ascii=False),
        "orders": json.dumps([{"field": "letter", "type": "asc"}], ensure_ascii=False),
        "returnInfos": json.dumps(return_infos, ensure_ascii=False),
        "articleType": 1,
        "level": 1,
    }
    payload = post_general_query(session, api_url, data)
    rows: list[dict[str, Any]] = []
    for item in payload.get("data", []):
        name = clean_faculty_name(item.get("title"))
        if not name:
            continue
        profile_url = first_http_url(item.get("cnUrl"))
        if not profile_url:
            profile_url = f"{config.directory_url}#{name}"
        rows.append(
            {
                "名录序号": len(rows) + 1,
                "姓名": name,
                "职称": unique_join([item.get("exField9", ""), item.get("exField1", ""), item.get("career", "")]),
                "教师主页链接": profile_url,
                "名录研究所": "复旦大学计算与智能创新学院",
                "所内职务": norm_text(item.get("exField1")) or norm_text(item.get("career")) or norm_text(item.get("exField9")),
                "邮箱": extract_email(norm_text(item.get("email"))),
                "电话": "",
                "地址": "",
                "个人主页": "",
                "主页研究所": "计算与智能创新学院",
                "研究方向": "",
                "个人简介摘要": unique_join([item.get("exField1", ""), item.get("career", "")])[:1200],
                "抓取状态": "待抓取" if valid_http_url(profile_url) and "#" not in profile_url else "成功-接口条目",
            }
        )

    parsed_rows: list[dict[str, Any]] = []
    for row in rows:
        if valid_http_url(row["教师主页链接"]) and "#" not in row["教师主页链接"]:
            parsed_rows.append(parse_generic_profile_detail(session, row))
            time.sleep(0.03)
        else:
            parsed_rows.append(row)
    return parsed_rows


SEU_DEPARTMENT_PAGE = "https://cse.seu.edu.cn/54820/list.htm"
SEU_TARGET_DEPARTMENTS = {
    "seu_cs": "计算机科学系",
    "seu_ce": "计算机工程系",
    "seu_imaging": "影像科学与技术系",
}


def clean_seu_name(text: str) -> str:
    text = norm_text(text).replace("(兼职)", "").replace("（兼职）", "").replace("(博士后)", "").replace("（博士后）", "")
    return clean_name(text)


def fetch_seu_department_map(session: requests.Session) -> dict[str, str]:
    soup = fetch(session, SEU_DEPARTMENT_PAGE, "utf-8")
    node = soup.select_one(".wp_articlecontent") or soup.select_one(".listcon")
    if node is None:
        return {}
    departments = {"计算机科学系", "计算机工程系", "影像科学与技术系"}
    current = ""
    result: dict[str, str] = {}
    for element in node.descendants:
        if isinstance(element, str):
            text = norm_text(element)
            if text in departments:
                current = text
        if getattr(element, "name", None) == "a" and element.get("href") and current:
            name = clean_seu_name(element.get_text(" ", strip=True))
            if name:
                result.setdefault(name, current)
    return result


def parse_seu_mentor_links(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    node = soup.select_one(".wp_articlecontent") or soup.select_one(".listcon")
    if node is None:
        return []
    rows_by_name: dict[str, dict[str, Any]] = {}
    current_section = ""
    for element in node.descendants:
        if isinstance(element, str):
            text = norm_text(element)
            if text in {"博士生导师", "硕士生导师"}:
                current_section = text
        if getattr(element, "name", None) != "a" or not element.get("href"):
            continue
        display_name = norm_text(element.get_text(" ", strip=True))
        name = clean_seu_name(display_name)
        if not name or name in {"首页", "师资队伍", "研究生导师", "教师按职称", "教师按研究方向", "教师按系别"}:
            continue
        href = urljoin(config.directory_url, element.get("href", ""))
        row = rows_by_name.setdefault(
            name,
            {
                "姓名": name,
                "显示姓名": display_name,
                "教师主页链接": href,
                "导师类型": [],
                "是否兼职": "是" if "兼职" in display_name else "",
            },
        )
        if current_section and current_section not in row["导师类型"]:
            row["导师类型"].append(current_section)
        if row["教师主页链接"] == config.directory_url and href:
            row["教师主页链接"] = href
    return list(rows_by_name.values())


def parse_seu_profile_detail(session: requests.Session, row: dict[str, Any], department_map: dict[str, str]) -> dict[str, Any]:
    detail = dict(row)
    profile_url = norm_text(row.get("教师主页链接"))
    parsed_profile = urlparse(profile_url)
    if parsed_profile.netloc.lower() in {"cs.seu.edu.cn", "cse.seu.edu.cn", "ai.seu.edu.cn"} and parsed_profile.path.rstrip("/") in {"", "/"}:
        detail["抓取状态"] = "无独立主页-学院首页"
        detail["职称"] = ""
        detail["邮箱"] = ""
        detail["电话"] = ""
        detail["地址"] = ""
        detail["个人主页"] = ""
        detail["研究方向"] = ""
        detail["个人简介摘要"] = ""
        return detail
    mentor_types = row.get("导师类型", [])
    if not isinstance(mentor_types, list):
        mentor_types = [mentor_types]
    detail["名录序号"] = row.get("名录序号", "")
    detail["职称"] = norm_text(row.get("职称"))
    detail["名录研究所"] = norm_text(row.get("官方系别")) or "院系待确认"
    detail["主页研究所"] = norm_text(row.get("官方系别")) or "院系待确认"
    detail["所内职务"] = unique_join(mentor_types + ["兼职" if row.get("是否兼职") else ""])
    if not valid_http_url(profile_url):
        detail["抓取状态"] = "无独立主页"
        detail["个人简介摘要"] = ""
        detail["研究方向"] = ""
        return detail

    try:
        soup = fetch(session, profile_url, "utf-8")
    except Exception as exc:  # noqa: BLE001
        detail["抓取状态"] = f"主页抓取失败：{exc}"
        detail["个人简介摘要"] = ""
        detail["研究方向"] = ""
        return detail

    content_text = get_article_text(soup)
    full_text = norm_text(soup.get_text(" ", strip=True))
    profile_text = content_text if len(content_text) > 50 else full_text
    for candidate in [row.get("显示姓名", ""), row.get("姓名", "")]:
        candidate = norm_text(candidate).replace("(兼职)", "").replace("（兼职）", "")
        name_index = profile_text.find(candidate)
        if candidate and name_index > 0:
            profile_text = profile_text[name_index:]
            break
    profile_text = strip_at_markers(profile_text, ["相关链接", "联系方式", "学院微信公众号", "Copyright"])
    department = extract_labeled_value(profile_text, "所在院系", ["研究方向", "电话", "邮箱", "职务", "个人简介"])
    if not department:
        department = department_map.get(norm_text(row.get("姓名")), "")
    title = extract_labeled_value(profile_text, "职称", ["所在院系", "研究方向", "电话", "邮箱", "职务", "个人简介"])
    direction = extract_labeled_value(profile_text, "研究方向", ["电话", "邮箱", "职务", "个人简介"])
    if title.startswith("教师按"):
        title = ""
    if direction.startswith(("电话", "邮箱", "职务", "个人简介", "研究方向")):
        direction = ""
    email = extract_labeled_value(profile_text, "邮箱", ["职务", "个人简介", "研究方向"])
    article_text = get_article_text(soup)
    if len(article_text) < 80:
        article_text = profile_text
    homepage = norm_text(row.get("个人主页"))
    if not homepage:
        homepage = infer_clean_personal_homepage(soup, profile_url)

    detail["抓取状态"] = "成功"
    detail["职称"] = title or norm_text(row.get("职称"))
    detail["官方系别"] = department
    detail["名录研究所"] = department or "院系待确认"
    detail["主页研究所"] = department or norm_text(row.get("官方系别")) or "院系待确认"
    detail["邮箱"] = extract_email(email) or extract_email(profile_text)
    detail["电话"] = extract_phone(profile_text)
    detail["地址"] = ""
    detail["个人主页"] = homepage
    has_empty_direction_label = "研究方向" in profile_text[:200] and not direction
    detail["研究方向"] = direction or ("" if has_empty_direction_label else direction_from_text(profile_text))
    detail["个人简介摘要"] = unique_join([detail["研究方向"], article_text])[:1200]
    return detail


def fetch_seu_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    department_map = fetch_seu_department_map(session)
    mentors = parse_seu_mentor_links(session, config)
    target_department = SEU_TARGET_DEPARTMENTS.get(config.key)
    rows: list[dict[str, Any]] = []
    for row in mentors:
        name = norm_text(row.get("姓名"))
        official_department = department_map.get(name, "")
        base_row = {
            "名录序号": "",
            "姓名": name,
            "显示姓名": row.get("显示姓名", ""),
            "职称": "",
            "教师主页链接": row.get("教师主页链接", ""),
            "名录研究所": official_department or "院系待确认",
            "所内职务": unique_join(row.get("导师类型", [])),
            "邮箱": "",
            "电话": "",
            "地址": "",
            "个人主页": "",
            "主页研究所": official_department or "院系待确认",
            "官方系别": official_department,
            "导师类型": row.get("导师类型", []),
            "研究方向": "",
            "个人简介摘要": "",
            "抓取状态": "待抓取",
            "是否兼职": row.get("是否兼职", ""),
        }
        detail = parse_seu_profile_detail(session, base_row, department_map)
        if target_department and norm_text(detail.get("官方系别")) != target_department:
            continue
        detail["名录序号"] = len(rows) + 1
        rows.append(detail)
        time.sleep(0.03)
    return rows


TONGJI_CS_CATEGORIES = [
    ("教授（研究员）", "https://cs.tongji.edu.cn/szdw/jsml_azc_/js_yjy_.htm"),
    ("副教授（副研究员）", "https://cs.tongji.edu.cn/szdw/jsml_azc_/fjs_fyjy_.htm"),
    ("预聘助理教授", "https://cs.tongji.edu.cn/szdw/jsml_azc_/ypzljs.htm"),
    ("讲师", "https://cs.tongji.edu.cn/szdw/jsml_azc_/js1.htm"),
    ("教辅系列", "https://cs.tongji.edu.cn/szdw/jsml_azc_/jfxl.htm"),
    ("兼职教授", "https://cs.tongji.edu.cn/szdw/jsml_azc_/jzjs.htm"),
]

TONGJI_SEE_BASE = "https://see.tongji.edu.cn/szdw1/jzyg"
TONGJI_SEE_LETTER_PAGES = ["A_G", "H_N", "O_T", "U_Z"]
TONGJI_SEE_CATEGORIES = [
    ("教授（研究员）", "jiaoshou"),
    ("副教授（副研究员）", "fjs"),
    ("讲师（助理教授）", "js"),
]
TONGJI_DETAIL_STOPS = [
    "姓名",
    "职称",
    "学科",
    "专业",
    "研究方向",
    "导师类型",
    "电子邮件",
    "联系电话",
    "通讯地址",
    "个人简介",
    "上一条",
    "下一条",
    "版权所有",
    "Copyright",
]


def tongji_label_variants(label: str) -> list[str]:
    variants = [label]
    if has_cjk(label):
        variants.append(" ".join(label))
    return variants


def extract_tongji_labeled_value(text: str, label: str, stops: list[str] | None = None) -> str:
    expanded_stops: list[str] = []
    for stop in stops or TONGJI_DETAIL_STOPS:
        expanded_stops.extend(tongji_label_variants(stop))
    for candidate in tongji_label_variants(label):
        value = extract_labeled_value(text, candidate, expanded_stops)
        if value:
            return value
    return ""


def collect_tongji_links(session: requests.Session, url: str, category: str) -> list[dict[str, Any]]:
    soup = fetch(session, url, "utf-8")
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for a in soup.select('a[href*="/info/"]'):
        href = urljoin(url, a.get("href", ""))
        name = clean_display_name(a.get("title", "") or a.get_text(" ", strip=True))
        if not name or len(name) < 2 or href in seen:
            continue
        seen.add(href)
        rows.append(
            {
                "姓名": name,
                "职称": category,
                "教师主页链接": href,
                "名录研究所": "",
                "所内职务": category,
                "邮箱": "",
                "电话": "",
                "地址": "",
                "个人主页": "",
                "主页研究所": "",
                "学科": "",
                "专业": "",
                "导师类型": "",
                "研究方向": "",
                "个人简介摘要": "",
                "抓取状态": "待抓取",
            }
        )
    return rows


def parse_tongji_profile_detail(session: requests.Session, row: dict[str, Any], config: TargetConfig) -> dict[str, Any]:
    detail = dict(row)
    profile_url = norm_text(row.get("教师主页链接"))
    if not valid_http_url(profile_url):
        detail["抓取状态"] = "无独立主页"
        return detail

    soup = fetch(session, profile_url, "utf-8")
    article_node = soup.select_one(".v_news_content, .wp_articlecontent, .wp_article_content, .article, .content")
    article_text = norm_text(article_node.get_text(" ", strip=True) if article_node else get_article_text(soup))
    main_node = soup.select_one(".main, .mainContent, .content, body")
    main_text = norm_text(main_node.get_text(" ", strip=True) if main_node else soup.get_text(" ", strip=True))

    has_structured_profile = config.key == "tongji_see" and bool(extract_tongji_labeled_value(main_text, "姓名"))
    title = extract_tongji_labeled_value(main_text, "职称") if has_structured_profile else ""
    if len(title) > 40 or (title and not re.search(r"教授|研究员|讲师|工程师|导师", title)):
        title = ""
    discipline = extract_tongji_labeled_value(main_text, "学科") if has_structured_profile else ""
    major = extract_tongji_labeled_value(main_text, "专业") if has_structured_profile else ""
    direction = (
        extract_tongji_labeled_value(main_text, "研究方向") if has_structured_profile else ""
    ) or direction_from_text(article_text)
    supervisor_type = extract_tongji_labeled_value(main_text, "导师类型") if has_structured_profile else ""
    email = (extract_tongji_labeled_value(main_text, "电子邮件") if has_structured_profile else "") or extract_email(article_text)
    phone = (extract_tongji_labeled_value(main_text, "联系电话") if has_structured_profile else "") or extract_phone(article_text)
    address = extract_tongji_labeled_value(main_text, "通讯地址") if has_structured_profile else ""

    detail["抓取状态"] = "成功"
    detail["职称"] = title or norm_text(row.get("职称"))
    detail["名录研究所"] = config.college_name
    detail["主页研究所"] = unique_join([discipline, major]) or config.college_name
    detail["学科"] = discipline
    detail["专业"] = major
    detail["导师类型"] = supervisor_type
    detail["邮箱"] = extract_email(email)
    detail["电话"] = phone
    detail["地址"] = address
    detail["个人主页"] = infer_clean_personal_homepage(soup, profile_url)
    detail["研究方向"] = direction
    detail["个人简介摘要"] = unique_join([direction, article_text])[:1200]
    return detail


def fetch_tongji_cs_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for category, url in TONGJI_CS_CATEGORIES:
        for row in collect_tongji_links(session, url, category):
            row["名录研究所"] = config.college_name
            row["主页研究所"] = config.college_name
            key = row["教师主页链接"] or row["姓名"]
            if key in merged:
                merged[key]["所内职务"] = unique_join([merged[key].get("所内职务", ""), category])
                merged[key]["职称"] = unique_join([merged[key].get("职称", ""), category])
            else:
                merged[key] = row

    rows: list[dict[str, Any]] = []
    for index, row in enumerate(merged.values(), start=1):
        row["名录序号"] = index
        try:
            rows.append(parse_tongji_profile_detail(session, row, config))
        except Exception as exc:
            row["抓取状态"] = f"失败：{exc}"
            rows.append(row)
        time.sleep(0.03)
    return rows


def fetch_tongji_see_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for category, category_slug in TONGJI_SEE_CATEGORIES:
        for letter_page in TONGJI_SEE_LETTER_PAGES:
            url = f"{TONGJI_SEE_BASE}/{category_slug}/{letter_page}.htm"
            for row in collect_tongji_links(session, url, category):
                row["名录研究所"] = config.college_name
                row["主页研究所"] = config.college_name
                row["拼音分组"] = letter_page.replace("_", "-")
                key = row["教师主页链接"] or row["姓名"]
                if key in merged:
                    merged[key]["所内职务"] = unique_join([merged[key].get("所内职务", ""), category])
                    merged[key]["职称"] = unique_join([merged[key].get("职称", ""), category])
                else:
                    merged[key] = row

    rows: list[dict[str, Any]] = []
    for index, row in enumerate(merged.values(), start=1):
        row["名录序号"] = index
        try:
            rows.append(parse_tongji_profile_detail(session, row, config))
        except Exception as exc:
            row["抓取状态"] = f"失败：{exc}"
            rows.append(row)
        time.sleep(0.03)
    return rows


ZJU_GENERIC_EMAILS = {"xwmaster@zju.edu.cn", "hgzly@zju.edu.cn"}
ZJU_CS_PDF_TABLE_HEADERS = {
    "序号",
    "所室机构",
    "姓名",
    "职  称",
    "学科专长及研究方向",
    "办公地点",
    "联系方式",
    "Email",
    "备注",
    "科研团队名称",
}
ZJU_CS_GENERIC_TEAM_NAMES = {"", "CCNT", "网安中心"}
ZJU_CS_TEAM_SUMMARY_TERMS = [
    "研究方向",
    "研究内容",
    "研究领域",
    "团队简介",
    "团队介绍",
    "实验室简介",
    "主要方向",
    "人工智能",
    "大模型",
    "具身智能",
    "机器人",
    "计算机视觉",
    "自动驾驶",
    "智能",
    "数据",
    "安全",
]


def normalize_zju_cs_pdf_url(raw_href: str) -> str:
    raw_href = norm_text(raw_href)
    if "/_upload/" in raw_href:
        return urljoin("http://www.cs.zju.edu.cn", raw_href[raw_href.index("/_upload/") :])
    return urljoin(ZJU_CS_TEAM_INFO_URL, raw_href)


def safe_pdf_cache_name(title: str, url: str) -> str:
    stem = re.sub(r"\.pdf$", "", norm_text(title), flags=re.I)
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", stem).strip("._ ")[:80]
    if not stem:
        stem = "attachment"
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()[:10]
    return f"{digest}_{stem}.pdf"


def fetch_zju_cs_pdf_links(session: requests.Session) -> list[dict[str, str]]:
    soup = fetch(session, ZJU_CS_TEAM_INFO_URL, "utf-8")
    links: list[dict[str, str]] = []
    seen: set[str] = set()
    for a in soup.select("a[href]"):
        raw_href = norm_text(a.get("href"))
        if ".pdf" not in raw_href.lower():
            continue
        url = normalize_zju_cs_pdf_url(raw_href)
        if url in seen:
            continue
        seen.add(url)
        title = norm_text(a.get_text(" ", strip=True)) or norm_text(a.get("title")) or Path(urlparse(url).path).name
        links.append({"title": title, "url": url})
    return links


def download_zju_cs_pdf(session: requests.Session, config: TargetConfig, link: dict[str, str]) -> Path:
    cache_dir = config.output_dir / "pdf_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    path = cache_dir / safe_pdf_cache_name(link["title"], link["url"])
    if path.exists() and path.stat().st_size > 1000:
        return path
    response = session.get(link["url"], timeout=45, headers=HEADERS)
    response.raise_for_status()
    path.write_bytes(response.content)
    return path


def extract_pdf_text(path: Path) -> tuple[str, str]:
    try:
        import fitz  # type: ignore[import-not-found]
    except ImportError:
        return "", "PyMuPDF未安装，无法抽取PDF文本"
    try:
        doc = fitz.open(path)
        text = "\n".join(page.get_text("text") for page in doc)
        status = f"成功：{doc.page_count}页，{len(text)}字"
        if len(norm_text(text)) < 50:
            status = f"文本层过少：{doc.page_count}页，可能是扫描件"
        return text, status
    except Exception as exc:  # noqa: BLE001
        return "", f"PDF解析失败：{type(exc).__name__}: {exc}"


def clean_zju_pdf_lines(text: str) -> list[str]:
    lines: list[str] = []
    for raw_line in text.splitlines():
        line = norm_text(raw_line)
        if not line or line in ZJU_CS_PDF_TABLE_HEADERS:
            continue
        if line.startswith("研究生导师信息库"):
            continue
        if re.search(r"第\s*\d+\s*页", line):
            continue
        lines.append(line)
    return lines


def looks_like_zju_pdf_name(value: str) -> bool:
    return bool(re.fullmatch(r"[\u4e00-\u9fff·]{2,8}", norm_text(value)))


def is_zju_supervisor_row_start(lines: list[str], index: int) -> bool:
    if index + 3 >= len(lines) or not re.fullmatch(r"\d{1,3}", lines[index]):
        return False
    institute = lines[index + 1]
    name = lines[index + 2]
    if "@" in institute or re.search(r"\d", institute) or len(institute) > 12:
        return False
    return looks_like_zju_pdf_name(name)


def is_zju_pdf_tail_field(value: str) -> bool:
    value = norm_text(value)
    if not value:
        return True
    if value in {"——", "--", "-", "—"}:
        return True
    if re.search(r"微信|QQ|电话|\d{5,}|@\w", value):
        return True
    location_pattern = r"楼|校区|园区|大楼|主楼|座|号|校外|西溪|玉泉|紫金港|之江实验室|良渚|曹主|曹东|则通|逸夫|智泉|\d+\s*室"
    if len(value) <= 35 and re.search(location_pattern, value):
        return True
    return False


def clean_zju_supervisor_direction(value: str) -> str:
    value = norm_text(value)
    location_tail_patterns = [
        r"\s+(?:曹主|曹东|曹光彪主楼|则通楼|老生仪楼|智泉大楼|逸夫工商楼|西溪校区|玉泉校区|紫金港校区|之江实验室|良渚实验室)[^，,、；;]{0,24}$",
        r"\s+\d{2,4}(?:[A-Za-z]|\（[^）]*\）|\([^)]*\))?$",
    ]
    changed = True
    while changed:
        changed = False
        for pattern in location_tail_patterns:
            new_value = re.sub(pattern, "", value).strip()
            if new_value != value:
                value = new_value
                changed = True
    return value


def parse_zju_supervisor_pdf(text: str) -> dict[str, dict[str, str]]:
    lines = clean_zju_pdf_lines(text)
    starts = [idx for idx in range(len(lines)) if is_zju_supervisor_row_start(lines, idx)]
    entries: dict[str, dict[str, str]] = {}
    for start_index, start in enumerate(starts):
        end = starts[start_index + 1] if start_index + 1 < len(starts) else len(lines)
        block = lines[start:end]
        if len(block) < 5:
            continue
        serial, institute, name, title = block[:4]
        rest = block[4:]
        email_index = next((idx for idx, line in enumerate(rest) if "@" in line), -1)
        email = ""
        pre_email = rest
        post_email: list[str] = []
        if email_index >= 0:
            email_text = "".join(rest[email_index : min(email_index + 2, len(rest))])
            email = extract_email(email_text) or norm_text(rest[email_index])
            pre_email = rest[:email_index]
            post_email = rest[email_index + 1 :]

        direction_end = len(pre_email)
        while direction_end > 0 and is_zju_pdf_tail_field(pre_email[direction_end - 1]):
            direction_end -= 1
        direction = clean_zju_supervisor_direction(" ".join(pre_email[:direction_end]))

        remark = ""
        team_lines: list[str] = []
        for idx, line in enumerate(post_email):
            if re.search(r"[博硕]导", line):
                remark = line
                team_lines = [item for item in post_email[idx + 1 :] if not is_zju_pdf_tail_field(item)]
                break
        team = norm_text(" ".join(team_lines))
        entries[name] = {
            "序号": serial,
            "所室机构": institute,
            "姓名": name,
            "职称": title,
            "研究方向": direction,
            "邮箱": email,
            "备注": remark,
            "科研团队名称": team,
        }
    return entries


def zju_team_tokens(value: str) -> list[str]:
    tokens = re.split(r"[;；、,，/| ]+", norm_text(value))
    return [
        token
        for token in tokens
        if len(token) >= 4 and token not in ZJU_CS_GENERIC_TEAM_NAMES and not re.fullmatch(r"[A-Za-z]{1,3}", token)
    ]


def zju_team_pdf_summary(title: str, text: str) -> str:
    lines = clean_zju_pdf_lines(text)
    important: list[str] = []
    for line in lines:
        if any(term in line for term in ZJU_CS_TEAM_SUMMARY_TERMS):
            important.append(line)
        if len(important) >= 10:
            break
    if not important:
        important = lines[:8]
    return unique_join([title, "；".join(important)])[:1000]


def zju_name_in_team_pdf(name: str, title: str, text: str) -> bool:
    name = norm_text(name)
    if not name:
        return False
    if name in title:
        return True
    search_text = text[:20000]
    if len(name) <= 2:
        pattern = rf"(?:{re.escape(name)}(?:教授|副教授|老师|研究员|博士|团队|课题组)|(?:负责人|团队成员|教师团队|团队负责人|导师)[^。；\n]{{0,80}}{re.escape(name)})"
        return bool(re.search(pattern, search_text))
    return name in search_text


def zju_team_name_in_pdf(team_name: str, title: str, text: str) -> bool:
    for token in zju_team_tokens(team_name):
        if token in title or token in text[:20000]:
            return True
    return False


def enrich_zju_cs_rows_with_pdf_info(
    session: requests.Session,
    config: TargetConfig,
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    try:
        pdf_links = fetch_zju_cs_pdf_links(session)
    except Exception as exc:  # noqa: BLE001
        print(f"zju_cs_pdf_links_error={type(exc).__name__}: {exc}")
        return rows

    pdf_docs: list[dict[str, Any]] = []
    for link in pdf_links:
        try:
            path = download_zju_cs_pdf(session, config, link)
            text, status = extract_pdf_text(path)
        except Exception as exc:  # noqa: BLE001
            path = Path("")
            text = ""
            status = f"PDF下载/解析失败：{type(exc).__name__}: {exc}"
        pdf_docs.append({**link, "path": str(path), "text": text, "status": status})

    supervisor_doc = next((doc for doc in pdf_docs if "全日制导师信息库" in doc["title"]), None)
    supervisor_entries = parse_zju_supervisor_pdf(supervisor_doc["text"]) if supervisor_doc else {}
    team_docs = [doc for doc in pdf_docs if doc is not supervisor_doc]

    enriched_rows: list[dict[str, Any]] = []
    supervisor_hits = 0
    team_hits = 0
    for row in rows:
        row = dict(row)
        name = norm_text(row.get("姓名"))
        pdf_sources: list[str] = []
        pdf_summaries: list[str] = []
        entry = supervisor_entries.get(name)
        if entry:
            supervisor_hits += 1
            pdf_direction = norm_text(entry.get("研究方向"))
            pdf_team = norm_text(entry.get("科研团队名称"))
            row["导师信息库研究方向"] = pdf_direction
            row["导师信息库团队"] = pdf_team
            row["研究方向"] = unique_join([row.get("研究方向", ""), f"导师信息库：{pdf_direction}" if pdf_direction else ""])
            row["职称"] = norm_text(row.get("职称")) or norm_text(entry.get("职称"))
            row["邮箱"] = norm_text(row.get("邮箱")) or norm_text(entry.get("邮箱"))
            row["名录研究所"] = norm_text(row.get("名录研究所")) or norm_text(entry.get("所室机构"))
            pdf_sources.append(f"{supervisor_doc['title']} {supervisor_doc['url']}" if supervisor_doc else "")
            pdf_summaries.append(
                unique_join(
                    [
                        f"2026全日制导师信息库：所室机构 {entry.get('所室机构')}",
                        f"职称 {entry.get('职称')}",
                        f"研究方向 {pdf_direction}",
                        f"科研团队 {pdf_team}" if pdf_team else "",
                    ],
                    "；",
                )
            )

        team_evidence: list[str] = []
        for doc in team_docs:
            text = norm_text(doc.get("text"))
            title = norm_text(doc.get("title"))
            if not text and not title:
                continue
            matched = zju_name_in_team_pdf(name, title, text)
            if not matched and entry:
                matched = zju_team_name_in_pdf(entry.get("科研团队名称", ""), title, text)
            if not matched:
                continue
            summary = zju_team_pdf_summary(title, text)
            if summary:
                team_evidence.append(summary)
                pdf_sources.append(f"{title} {doc.get('url')}")
        if team_evidence:
            team_hits += 1
            row["团队PDF证据"] = unique_join(team_evidence[:3])[:1800]
            pdf_summaries.append(f"导师团队PDF：{row['团队PDF证据']}")
        else:
            row.setdefault("团队PDF证据", "")

        row["导师信息库PDF"] = unique_join(pdf_sources[:6])
        if pdf_summaries:
            row["个人简介摘要"] = unique_join([row.get("个人简介摘要", ""), *pdf_summaries])[:2200]
        row.setdefault("导师信息库研究方向", "")
        row.setdefault("导师信息库团队", "")
        enriched_rows.append(row)

    print(
        "zju_cs_pdf_supplement="
        f"links:{len(pdf_links)}, supervisor_entries:{len(supervisor_entries)}, "
        f"supervisor_hits:{supervisor_hits}, team_hits:{team_hits}"
    )
    return enriched_rows


def clean_zju_name(text: str) -> str:
    return clean_name(norm_text(text).replace("\u3000", " "))


def clean_zju_emails(text: str) -> str:
    emails = [
        email
        for email in extract_email(text).split("; ")
        if email and email.lower() not in ZJU_GENERIC_EMAILS
    ]
    return unique_join(emails)


def extract_zju_research_direction(text: str) -> str:
    source = norm_text(text)
    footer_stops = [
        "研究主题",
        "个人简介",
        "论文发表",
        "科研项目",
        "人才培养",
        "出版著作",
        "专利成果",
        "学术服务",
        "教学工作",
        "版权所有",
        "Copyright",
        "通讯地址",
        "邮编",
        "技术支持",
        "相关链接",
    ]
    direction = extract_labeled_value(
        source,
        "研究方向",
        footer_stops,
    )
    topic = extract_labeled_value(source, "研究主题", ["个人简介", "版权所有", "Copyright", "通讯地址", "邮编", "技术支持", "相关链接"])
    if direction:
        return unique_join([direction, f"研究主题：{topic}" if topic else ""])
    return direction_from_text(source)


def extract_zju_title(text: str, name: str) -> str:
    source = norm_text(text)
    for label in ["职位", "职称"]:
        value = extract_labeled_value(source, label, ["Email", "邮箱", "个人主页", "个人介绍", "研究方向", "单位", "职务", "电话"])
        if value:
            return value
    title_match = re.search(
        r"((?:求是讲席教授、)?(?:长聘教授|教授|副教授|助理教授|研究员|副研究员|讲师|高级工程师|工程师)(?:[、/| ]*(?:博士生导师|硕士生导师|研究员|副研究员|教授|副教授))*)",
        source,
    )
    if title_match:
        return norm_text(title_match.group(1))
    if name:
        after_name = source.split(name, 1)[1] if name in source else source
        after_name = strip_at_markers(after_name, ["单位", "职务", "电话", "邮箱", "地址", "研究方向", "个人简介"])
        title_tokens = [
            token
            for token in re.split(r"\s*[|/]\s*|\s+", after_name)
            if token
            and token not in {"博士", "硕士", "链接", "个人主页", "更新时间", "总访问量"}
            and not token.startswith(("首页", "学校概况", "浙大服务", "关于主页"))
        ]
        title = unique_join(title_tokens[:4])
        if title:
            return title
    return ""


def resolve_zju_redirect(session: requests.Session, url: str) -> str:
    parsed = urlparse(url)
    if parsed.netloc.lower() != "ai.zju.edu.cn" or parsed.path != "/_redirect":
        return url
    try:
        response = session.get(url, timeout=8, headers=HEADERS, allow_redirects=False)
        if response.status_code in {301, 302, 303, 307, 308} and response.headers.get("Location"):
            return urljoin(url, response.headers["Location"])
    except Exception:
        pass
    return url


def plausible_zju_title(value: str) -> bool:
    value = norm_text(value)
    if not value:
        return False
    if len(value) > 80:
        return False
    if re.search(r"\b(home|research|publication|toggle|navigation|university|zhejiang)\b", value, flags=re.I):
        return False
    return bool(re.search(r"教授|研究员|讲师|工程师|导师|院士|PI|Professor|Researcher|Lecturer|Engineer", value, flags=re.I))


def parse_zju_profile_detail(session: requests.Session, row: dict[str, Any], homepage_as_personal: bool = False) -> dict[str, Any]:
    detail = dict(row)
    profile_url = norm_text(row.get("教师主页链接"))
    if not valid_http_url(profile_url):
        detail["抓取状态"] = "无独立主页"
        return detail

    resolved_profile_url = resolve_zju_redirect(session, profile_url)
    last_error: Exception | None = None
    try:
        for attempt in range(2):
            try:
                response = session.get(resolved_profile_url, timeout=20, headers=HEADERS)
                response.raise_for_status()
                response.encoding = "utf-8"
                final_url = response.url
                soup = BeautifulSoup(response.text, "html.parser")
                break
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                if attempt == 1:
                    raise
                time.sleep(0.2)
    except Exception as exc:  # noqa: BLE001
        detail["抓取状态"] = f"主页抓取失败：{last_error or exc}"
        return detail

    full_text = norm_text(soup.get_text(" ", strip=True))
    article_text = get_article_text(soup)
    if len(article_text) < 80:
        article_text = full_text

    name = norm_text(row.get("姓名"))
    title = extract_zju_title(full_text, name)
    if not plausible_zju_title(title):
        title = norm_text(row.get("职称"))
    email = clean_zju_emails(full_text)
    phone = extract_phone(full_text)
    direction = extract_zju_research_direction(full_text)
    homepage = norm_text(row.get("个人主页"))
    if not homepage:
        labeled_homepage = extract_labeled_value(
            full_text,
            "个人主页",
            ["个人介绍", "研究方向", "个人简介", "Email", "邮箱", "地址", "电话"],
        )
        homepage = first_http_url(labeled_homepage)
    if not homepage:
        homepage = infer_clean_personal_homepage(soup, profile_url)
    if homepage_as_personal and not homepage:
        homepage = final_url or profile_url

    detail["抓取状态"] = "成功"
    detail["职称"] = title or norm_text(row.get("职称"))
    detail["邮箱"] = norm_text(row.get("邮箱")) or email
    detail["电话"] = norm_text(row.get("电话")) or phone
    detail["个人主页"] = homepage
    detail["研究方向"] = norm_text(row.get("研究方向")) or direction
    detail["个人简介摘要"] = unique_join([row.get("个人简介摘要", ""), detail["研究方向"], article_text])[:1200]
    return detail


def fetch_zju_cs_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for block in soup.select("li.wp_sublist"):
        title_node = block.select_one("h3.sublist_title .subcolumn-name")
        institute = norm_text(title_node.get_text(" ", strip=True) if title_node else "")
        if not institute:
            continue
        for a in block.select("ul.zuzhi a[title][href]"):
            name = clean_zju_name(a.get("title") or a.get_text(" ", strip=True))
            if not name or name in {"首页", "教师名录", "浏览更多"}:
                continue
            raw_href = norm_text(a.get("href"))
            href = "" if raw_href.lower().startswith("javascript:") else urljoin(config.directory_url, raw_href)
            key = (name, href or institute)
            if key in seen:
                continue
            seen.add(key)
            base_row = {
                "名录序号": len(rows) + 1,
                "姓名": name,
                "职称": "",
                "教师主页链接": href or f"{config.directory_url}#{name}",
                "名录研究所": institute,
                "所内职务": "教师",
                "邮箱": "",
                "电话": "",
                "地址": "",
                "个人主页": href if valid_http_url(href) and urlparse(href).netloc.lower() in {"person.zju.edu.cn", "mypage.zju.edu.cn"} else "",
                "主页研究所": institute,
                "研究方向": "",
                "个人简介摘要": "",
                "抓取状态": "待抓取" if valid_http_url(href) else "无独立主页",
            }
            detail = (
                parse_zju_profile_detail(session, base_row, homepage_as_personal=True)
                if valid_http_url(href)
                else base_row
            )
            detail["名录序号"] = len(rows) + 1
            rows.append(detail)
            time.sleep(0.02)
    return enrich_zju_cs_rows_with_pdf_info(session, config, rows)


def fetch_zju_ai_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for block in soup.select(".col_news_list li.wp_sublist"):
        title_node = block.select_one("h3.sublist_title .subcolumn-name")
        institute = norm_text(title_node.get_text(" ", strip=True) if title_node else "")
        if not institute:
            continue
        for a in block.select(".news_title a[title][href]"):
            name = clean_zju_name(a.get("title") or a.get_text(" ", strip=True))
            raw_href = norm_text(a.get("href"))
            href = "" if raw_href.lower().startswith("javascript:") else urljoin(config.directory_url, raw_href)
            if not name or not href:
                continue
            key = (name, href)
            if key in seen:
                continue
            seen.add(key)
            base_row = {
                "名录序号": len(rows) + 1,
                "姓名": name,
                "职称": "",
                "教师主页链接": href,
                "名录研究所": institute,
                "官方系别": institute,
                "所内职务": "教师",
                "邮箱": "",
                "电话": "",
                "地址": "",
                "个人主页": "",
                "主页研究所": institute,
                "研究方向": "",
                "个人简介摘要": "",
                "抓取状态": "待抓取",
            }
            detail = parse_zju_profile_detail(session, base_row, homepage_as_personal=True)
            detail["名录序号"] = len(rows) + 1
            detail["主页研究所"] = norm_text(detail.get("主页研究所")) or institute
            detail["官方系别"] = norm_text(detail.get("官方系别")) or institute
            rows.append(detail)
            time.sleep(0.02)
    return rows


USTC_EMPTY_TEXTS = {"", "暂无内容", "【查看更多】", "查看更多", "-", "--", "."}
USTC_LABELS = [
    "教师姓名",
    "教师英文名称",
    "电子邮箱",
    "邮箱",
    "职务",
    "职称",
    "学历",
    "学位",
    "所属院系",
    "学科",
    "办公地点",
    "联系方式",
    "毕业院校",
]
USTC_SECTION_HEADINGS = {
    "首页",
    "科学研究",
    "教学信息",
    "获奖信息",
    "招生信息",
    "研究领域",
    "论文成果",
    "专利",
    "著作成果",
    "科研项目",
    "教学研究",
    "教学资源",
    "授课信息",
    "教学成果",
    "学生信息",
    "我的相册",
    "教师博客",
    "基本信息",
    "个人信息",
    "Personal Information",
    "个人简介",
    "教育经历",
    "Education Background",
    "工作经历",
    "Work Experience",
    "社会兼职",
    "Social Affiliations",
    "研究方向",
    "Research Focus",
    "团队成员",
    "其他联系方式",
    "联系方式",
    "导航",
    "语种",
    "更多",
}
USTC_TITLE_CANDIDATES = [
    "讲席教授",
    "预聘制副教授",
    "预聘副教授",
    "特任教授",
    "副教授",
    "助理教授",
    "教授",
    "副研究员",
    "研究员",
    "讲师",
    "博士后",
]


def is_ustc_empty_text(value: Any) -> bool:
    text = norm_text(value).replace("\u200b", "")
    return text in USTC_EMPTY_TEXTS


def looks_like_ustc_encrypted_text(value: Any) -> bool:
    return bool(re.fullmatch(r"[0-9a-f]{80,}", norm_text(value).lower()))


def clean_ustc_section_text(value: Any) -> str:
    text = norm_text(value).replace("\u200b", "").replace("【查看更多】", "")
    text = re.sub(r"\b[PM]\s*同专业[博硕]导\b", " ", text)
    text = norm_text(text)
    if is_ustc_empty_text(text) or looks_like_ustc_encrypted_text(text):
        return ""
    return text


def strip_ustc_heading_prefix(text: str, heading: str) -> str:
    text = clean_ustc_section_text(text)
    heading = norm_text(heading)
    if heading and text.startswith(heading):
        text = norm_text(text[len(heading) :])
    return clean_ustc_section_text(text)


def add_ustc_section(sections: dict[str, str], heading: Any, content: Any) -> None:
    heading_text = norm_text(heading)
    content_text = clean_ustc_section_text(content)
    if not heading_text or not content_text:
        return
    sections[heading_text] = unique_join([sections.get(heading_text, ""), content_text])


def line_starts_with_ustc_label(line: str) -> bool:
    return any(line.startswith(f"{label}：") or line.startswith(f"{label}:") for label in USTC_LABELS)


def ustc_text_lines(soup: BeautifulSoup) -> list[str]:
    return [norm_text(line) for line in soup.get_text("\n", strip=True).splitlines() if norm_text(line)]


def collect_next_ustc_label_value(lines: list[str], index: int) -> str:
    values: list[str] = []
    for line in lines[index + 1 : index + 6]:
        if line_starts_with_ustc_label(line) or line in USTC_SECTION_HEADINGS:
            break
        if is_ustc_empty_text(line) or looks_like_ustc_encrypted_text(line):
            break
        values.append(line)
    return unique_join(values, " ")


def extract_ustc_labeled_values(soup: BeautifulSoup) -> dict[str, str]:
    labels: dict[str, str] = {}
    lines = ustc_text_lines(soup)
    for index, line in enumerate(lines):
        for label in USTC_LABELS:
            value = ""
            if line.startswith(f"{label}："):
                value = norm_text(line.split("：", 1)[1])
            elif line.startswith(f"{label}:"):
                value = norm_text(line.split(":", 1)[1])
            else:
                continue
            if is_ustc_empty_text(value) or looks_like_ustc_encrypted_text(value):
                value = collect_next_ustc_label_value(lines, index)
            if value and not looks_like_ustc_encrypted_text(value):
                labels[label] = unique_join([labels.get(label, ""), value])
    return labels


def decrypt_ustc_encrypted_fields(session: requests.Session, soup: BeautifulSoup, profile_url: str) -> str:
    scripts_text = "\n".join(script.get_text(" ", strip=True) for script in soup.select("script"))
    mode_match = re.search(r"_tsites_com_view_mode_type_\s*=\s*(\d+)", scripts_text)
    mode = mode_match.group(1) if mode_match else "8"
    values: list[str] = []
    api_url = urljoin(profile_url, "/system/resource/tsites/tsitesencrypt.jsp")
    for span in soup.select("span[_tsites_encrypt_field]"):
        encrypted = norm_text(span.get_text(" ", strip=True))
        field_id = norm_text(span.get("id"))
        if not encrypted or not field_id:
            continue
        try:
            response = session.get(
                api_url,
                params={"id": field_id, "content": encrypted, "mode": mode},
                timeout=30,
                headers=HEADERS,
            )
            response.raise_for_status()
            content = response.json().get("content", "")
            decoded = BeautifulSoup(str(content), "html.parser").get_text(" ", strip=True)
            if decoded:
                values.append(decoded)
        except Exception:
            continue
    return unique_join(values)


def extract_ustc_sections(soup: BeautifulSoup) -> dict[str, str]:
    sections: dict[str, str] = {}

    for block in soup.select(".CurrencyBox"):
        heading_node = block.select_one(".CurrencyBit")
        heading = norm_text(heading_node.get_text(" ", strip=True) if heading_node else "")
        content_node = block.select_one(".EducationExperience, .text, .team-con")
        content = norm_text(content_node.get_text(" ", strip=True) if content_node else "")
        add_ustc_section(sections, heading, content)

    for block in soup.select(".box-2"):
        heading_node = block.select_one(".index-title")
        heading = norm_text(heading_node.get_text(" ", strip=True) if heading_node else "")
        content_node = block.select_one(".text, .team-con")
        content = norm_text(content_node.get_text(" ", strip=True) if content_node else "")
        add_ustc_section(sections, heading, content)

    for container in soup.select(".exper_cont"):
        headings = [norm_text(node.get_text(" ", strip=True)) for node in container.select(".exper_title")]
        contents = [norm_text(node.get_text(" ", strip=True)) for node in container.select(".tabs_cont_1")]
        for heading, content in zip(headings, contents, strict=False):
            add_ustc_section(sections, heading, strip_ustc_heading_prefix(content, heading))

    for container in soup.select(".TabbedPanels"):
        headings = [norm_text(node.get_text(" ", strip=True)) for node in container.select(".TabbedPanelsTab")]
        contents = [norm_text(node.get_text(" ", strip=True)) for node in container.select(".TabbedPanelsContent")]
        for heading, content in zip(headings, contents, strict=False):
            add_ustc_section(sections, heading, content)

    for block in soup.select(".rightcon"):
        heading_node = block.find(["h1", "h2", "h3", "h4"])
        heading = norm_text(heading_node.get_text(" ", strip=True) if heading_node else "")
        add_ustc_section(sections, heading, strip_ustc_heading_prefix(block.get_text(" ", strip=True), heading))

    intro_node = soup.select_one(".te_one_f")
    if intro_node:
        add_ustc_section(sections, "个人简介", intro_node.get_text(" ", strip=True))

    return sections


def extract_ustc_section_from_lines(lines: list[str], heading: str) -> str:
    candidates: list[str] = []
    for index, line in enumerate(lines):
        if line != heading:
            continue
        bits: list[str] = []
        for next_line in lines[index + 1 : index + 16]:
            if next_line in USTC_SECTION_HEADINGS or line_starts_with_ustc_label(next_line):
                break
            if is_ustc_empty_text(next_line) or looks_like_ustc_encrypted_text(next_line):
                continue
            if next_line in {"当前位置:", "中文主页", "版权所有 ©2020 中国科学技术大学"}:
                break
            bits.append(next_line)
        candidate = unique_join(bits, " ")
        if candidate:
            candidates.append(candidate)
    return max(candidates, key=len) if candidates else ""


def ustc_direction_from_text(text: str) -> str:
    text = norm_text(text)
    patterns = [
        r"(主要研究方向包括[^。；;\n]{4,220})",
        r"(研究方向包括[^。；;\n]{4,220})",
        r"(主要从事[^。；;\n]{4,220})",
        r"(长期围绕[^。；;\n]{4,220})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            return norm_text(match.group(1))
    fallback = direction_from_text(text)
    if any(token in fallback for token in ["暂无内容", "团队成员", "版权所有", "地址："]):
        return ""
    return fallback


def extract_ustc_title(text: Any) -> str:
    source = norm_text(text)
    for title in USTC_TITLE_CANDIDATES:
        if title in source:
            return title
    return ""


def infer_ustc_title(lines: list[str], labels: dict[str, str]) -> str:
    for value in [labels.get("职称", ""), labels.get("职务", "")]:
        title = extract_ustc_title(value)
        if title:
            return title
    for line in lines[:80]:
        title = extract_ustc_title(line)
        if title and "导师" not in line:
            return title
    return ""


def extract_ustc_external_urls(full_text: str, profile_url: str) -> str:
    urls = extract_urls(full_text).split("; ") if extract_urls(full_text) else []
    profile_host = urlparse(profile_url).netloc.lower()
    external_urls: list[str] = []
    for url in urls:
        parsed = urlparse(url)
        host = parsed.netloc.lower()
        if not host or host == profile_host:
            continue
        if host in GENERIC_COMMON_HOMEPAGE_HOSTS or host in GENERIC_NON_HOMEPAGE_HOSTS or "dblp." in host:
            continue
        external_urls.append(url)
    return unique_join(external_urls)


def parse_ustc_profile_detail(session: requests.Session, row: dict[str, Any]) -> dict[str, Any]:
    detail = dict(row)
    profile_url = norm_text(row.get("教师主页链接"))
    if not valid_http_url(profile_url):
        detail["抓取状态"] = "无独立主页"
        return detail

    try:
        soup = fetch(session, profile_url, "utf-8")
    except Exception as exc:  # noqa: BLE001
        detail["抓取状态"] = f"主页抓取失败：{exc}"
        return detail

    lines = ustc_text_lines(soup)
    labels = extract_ustc_labeled_values(soup)
    sections = extract_ustc_sections(soup)
    full_text = norm_text(soup.get_text(" ", strip=True))
    decoded_email = decrypt_ustc_encrypted_fields(session, soup, profile_url)
    raw_email = extract_email(full_text)
    email_label = unique_join([labels.get("电子邮箱", ""), labels.get("邮箱", ""), labels.get("联系方式", "")])

    direction = (
        sections.get("研究方向", "")
        or sections.get("Research Focus", "")
        or extract_ustc_section_from_lines(lines, "研究方向")
        or ustc_direction_from_text(sections.get("个人简介", ""))
        or ustc_direction_from_text(full_text)
    )
    summary = (
        sections.get("个人简介", "")
        or extract_ustc_section_from_lines(lines, "个人简介")
        or unique_join([direction, sections.get("科学研究", "")])
    )
    if not summary:
        summary = unique_join([labels.get("职务", ""), labels.get("职称", ""), direction])

    personal_homepage = unique_join(
        [
            infer_clean_personal_homepage(soup, profile_url),
            extract_ustc_external_urls(full_text, profile_url),
        ]
    )
    if not personal_homepage:
        personal_homepage = profile_url

    department = norm_text(labels.get("所属院系")) or "人工智能与数据科学学院"
    detail.update(
        {
            "职称": norm_text(row.get("职称")) or infer_ustc_title(lines, labels),
            "所内职务": norm_text(labels.get("职务")) or norm_text(row.get("所内职务")),
            "主页研究所": department,
            "官方系别": department,
            "邮箱": unique_join([row.get("邮箱", ""), decoded_email, raw_email, email_label]),
            "电话": norm_text(row.get("电话")) or extract_phone(full_text),
            "地址": norm_text(row.get("地址")) or norm_text(labels.get("办公地点")),
            "个人主页": personal_homepage,
            "研究方向": clean_ustc_section_text(direction),
            "个人简介摘要": clean_ustc_section_text(summary)[:1200],
            "抓取状态": "成功",
        }
    )
    return detail


def fetch_ustc_ai_ds_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for a in soup.select(".tc_list a[href]"):
        name_node = a.select_one(".tc_name")
        institute_node = a.select_one(".yx_name")
        name = clean_faculty_name(name_node.get_text(" ", strip=True) if name_node else a.get_text(" ", strip=True))
        href = urljoin(config.directory_url, norm_text(a.get("href")))
        if not name or "faculty.ustc.edu.cn" not in urlparse(href).netloc:
            continue
        key = (name, href)
        if key in seen:
            continue
        seen.add(key)
        institute = norm_text(institute_node.get_text(" ", strip=True) if institute_node else "") or config.college_name
        base_row = {
            "名录序号": len(rows) + 1,
            "姓名": name,
            "职称": "",
            "教师主页链接": href,
            "名录研究所": institute,
            "官方系别": institute,
            "所内职务": "教师",
            "邮箱": "",
            "电话": "",
            "地址": "",
            "个人主页": "",
            "主页研究所": institute,
            "研究方向": "",
            "个人简介摘要": "",
            "抓取状态": "待抓取",
        }
        rows.append(parse_ustc_profile_detail(session, base_row))
        time.sleep(0.03)
    return rows


def parse_zju_cse_profile_detail(session: requests.Session, row: dict[str, Any]) -> dict[str, Any]:
    detail = parse_zju_profile_detail(session, row)
    detail["主页研究所"] = norm_text(detail.get("主页研究所")) or norm_text(row.get("主页研究所")) or norm_text(row.get("名录研究所"))
    return detail


def fetch_zju_cse_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    current_institute = ""
    for node in soup.select("h3.research-name, li.teacher-item a[href]"):
        if node.name == "h3":
            current_institute = norm_text(node.get_text(" ", strip=True))
            continue
        name = clean_zju_name(node.get_text(" ", strip=True))
        raw_href = norm_text(node.get("href"))
        href = "" if raw_href.lower().startswith("javascript:") else urljoin(config.directory_url, raw_href)
        if not name or not href:
            continue
        key = (name, href)
        if key in seen:
            continue
        seen.add(key)
        base_row = {
            "名录序号": len(rows) + 1,
            "姓名": name,
            "职称": "",
            "教师主页链接": href,
            "名录研究所": current_institute,
            "所内职务": "教师",
            "邮箱": "",
            "电话": "",
            "地址": "",
            "个人主页": "",
            "主页研究所": current_institute,
            "研究方向": "",
            "个人简介摘要": "",
            "抓取状态": "待抓取",
        }
        detail = parse_zju_cse_profile_detail(session, base_row)
        detail["名录序号"] = len(rows) + 1
        rows.append(detail)
        time.sleep(0.02)
    return rows


def fetch_zju_uiuc_directory(session: requests.Session, config: TargetConfig) -> list[dict[str, Any]]:
    soup = fetch(session, config.directory_url, "utf-8")
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in soup.select(".item"):
        link = item.select_one("a[href]")
        if link is None:
            continue
        name = norm_text(link.get_text(" ", strip=True))
        href = urljoin(config.directory_url, link.get("href", ""))
        if not name or href in seen:
            continue
        seen.add(href)
        parts = [norm_text(part) for part in item.stripped_strings]
        rest = parts[parts.index(name) + 1 :] if name in parts else parts[1:]
        title_parts: list[str] = []
        direction_parts: list[str] = []
        for part in rest:
            if part == "/":
                continue
            if any(token in part for token in ["教授", "研究员", "院长", "首席教授", "讲席教授"]):
                title_parts.append(part)
            else:
                direction_parts.append(part)
        title = unique_join(title_parts)
        direction = unique_join(direction_parts)
        base_row = {
            "名录序号": len(rows) + 1,
            "姓名": name,
            "职称": title,
            "教师主页链接": href,
            "名录研究所": "浙江大学-伊利诺伊大学厄巴纳香槟校区联合学院",
            "所内职务": title,
            "邮箱": "",
            "电话": "",
            "地址": "",
            "个人主页": "",
            "主页研究所": "ZJUI",
            "研究方向": direction,
            "个人简介摘要": direction,
            "抓取状态": "待抓取",
        }
        detail = parse_zju_profile_detail(session, base_row)
        detail["名录序号"] = len(rows) + 1
        rows.append(detail)
        time.sleep(0.02)
    return rows


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill("solid", fgColor="D9EAD3")
    mid_fill = PatternFill("solid", fgColor="FFF2CC")
    low_fill = PatternFill("solid", fgColor="F4CCCC")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        widths = {
            "A": 10,
            "B": 12,
            "C": 14,
            "D": 12,
            "E": 12,
            "F": 10,
            "G": 24,
            "H": 48,
            "I": 56,
            "J": 24,
            "K": 24,
            "L": 26,
            "M": 26,
            "N": 18,
            "O": 24,
            "P": 34,
            "Q": 42,
            "R": 14,
            "S": 60,
        }
        for idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = widths.get(col_letter, 20)

        if ws.max_row >= 2 and "推荐等级" in [cell.value for cell in ws[1]]:
            level_col = [cell.value for cell in ws[1]].index("推荐等级") + 1
            col_letter = get_column_letter(level_col)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"强烈建议"'], fill=high_fill),
            )
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"可以考虑"'], fill=mid_fill),
            )
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"暂不优先"'], fill=low_fill),
            )

    wb.save(path)


def build_workbook(config: TargetConfig, rows: list[dict[str, Any]]) -> None:
    run_context = create_run_context("first_pass", config.key)
    scored_rows: list[dict[str, Any]] = []
    for row in rows:
        row = ensure_teacher_identity(config.school_slug, config.college_slug, row)
        row.update(score_first_pass_teacher(row))
        scored_rows.append(row)

    df = pd.DataFrame(scored_rows)
    ordered_cols = [
        "名录序号",
        "姓名",
        TEACHER_ID_COLUMN,
        IDENTITY_CONFIDENCE_COLUMN,
        "套磁情况",
        "职称",
        "推荐等级",
        "是否建议套磁",
        "匹配分",
        "命中关键词",
        "推荐理由",
        "显式核心锚点",
        "评分规则版本",
        "官方证据分",
        "DBLP证据分",
        "arXiv证据分",
        "网页证据分",
        "WebSearch证据分",
        "评分警告",
        "研究方向",
        "导师信息库研究方向",
        "导师信息库团队",
        "团队PDF证据",
        "去重备注",
        "名录研究所",
        "主页研究所",
        "官方系别",
        "所内职务",
        "是否兼职",
        "邮箱",
        "电话",
        "地址",
        "个人主页",
        "教师主页链接",
        "导师信息库PDF",
        "抓取状态",
        "个人简介摘要",
    ]
    for col in ordered_cols:
        if col not in df.columns:
            df[col] = ""
    df = df[ordered_cols]
    df = df.map(clean_excel_cell)
    df = apply_contact_statuses(df, config.school_slug, config.college_slug)

    priority_order = {"强烈建议": 0, "可以考虑": 1, "暂不优先": 2}
    df["_order"] = df["推荐等级"].map(priority_order).fillna(9)
    df = df.sort_values(["_order", "匹配分", "名录序号"], ascending=[True, False, True]).drop(columns=["_order"])
    priority_df = df[df["是否建议套磁"] == "是"].copy()

    source_rows = [
        {"项目": "抓取日期", "内容": TODAY},
        {"项目": "学校", "内容": config.school_name},
        {"项目": "学院", "内容": config.college_name},
        {"项目": "教师名录来源", "内容": config.directory_url},
        {"项目": "输出目录", "内容": str(config.output_dir)},
        {"项目": "匹配依据", "内容": RESUME_MATCH_CONTEXT},
        {
            "项目": "标注说明",
            "内容": "第一阶段匹配分来自学生画像中的关键词权重，以及教师主页/简介中的方向证据；后续 DBLP/arXiv/web 阶段会继续更新证据。",
        },
    ]
    source_rows.extend(context_source_rows(run_context))
    if config.key == "zju_cs":
        source_rows.append({"项目": "导师团队信息库PDF来源", "内容": ZJU_CS_TEAM_INFO_URL})
    source_df = pd.DataFrame(source_rows)
    source_df = source_df.map(clean_excel_cell)

    config.output_dir.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(config.output_path, engine="openpyxl") as writer:
        priority_df.to_excel(writer, sheet_name="优先套磁名单", index=False)
        df.to_excel(writer, sheet_name="全量教师名录", index=False)
        source_df.to_excel(writer, sheet_name="匹配依据", index=False)
    style_workbook(config.output_path)
    migrate_workbook(config.output_path, {})
    write_stage_manifest(config.output_dir, run_context)

    print(f"target={config.key}")
    print(f"rows={len(df)}")
    print(f"priority={len(priority_df)}")
    print(f"output={config.output_path.resolve()}")
    print(df["推荐等级"].value_counts().to_string())


def fetch_target_rows(config: TargetConfig) -> list[dict[str, Any]]:
    session = requests.Session()
    session.trust_env = False
    if config.key == "sjtu_cs":
        rows = fetch_sjtu_cs_directory(session, config)
    elif config.key == "sjtu_ai":
        rows = fetch_sjtu_ai_directory(session, config)
    elif config.key == "nju_cs":
        rows = fetch_nju_cs_directory(session, config)
    elif config.key == "nju_ai":
        rows = fetch_nju_ai_directory(session, config)
    elif config.key == "ruc_gsai":
        rows = fetch_ruc_gsai_directory(session, config)
    elif config.key == "ruc_ssai":
        rows = fetch_ruc_ssai_directory(session, config)
    elif config.key == "ruc_info":
        rows = fetch_ruc_info_directory(session, config)
    elif config.key == "nju_ra":
        rows = fetch_nju_ra_directory(session, config)
    elif config.key == "nju_is":
        rows = fetch_nju_is_directory(session, config)
    elif config.key == "nju_ic":
        rows = fetch_nju_ic_directory(session, config)
    elif config.key == "fudan_ciram":
        rows = fetch_fudan_ciram_directory(session, config)
    elif config.key == "fudan_ai":
        rows = fetch_fudan_ai_directory(session, config)
    elif config.key in {"seu_cs", "seu_ce", "seu_imaging"}:
        rows = fetch_seu_directory(session, config)
    elif config.key == "tongji_cs":
        rows = fetch_tongji_cs_directory(session, config)
    elif config.key == "tongji_see":
        rows = fetch_tongji_see_directory(session, config)
    elif config.key == "zju_cs":
        rows = fetch_zju_cs_directory(session, config)
    elif config.key == "zju_ai":
        rows = fetch_zju_ai_directory(session, config)
    elif config.key == "ustc_ai_ds":
        rows = fetch_ustc_ai_ds_directory(session, config)
    elif config.key == "zju_uiuc":
        rows = fetch_zju_uiuc_directory(session, config)
    elif config.key == "zju_cse":
        rows = fetch_zju_cse_directory(session, config)
    else:
        raise ValueError(config.key)
    return deduplicate_faculty_rows(rows, config)


def build_target(config: TargetConfig) -> None:
    rows = fetch_target_rows(config)
    build_workbook(config, rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build first-pass faculty matching workbooks.")
    parser.add_argument(
        "targets",
        nargs="*",
        default=list(TARGETS),
        help=f"Targets to build. Available: {', '.join(TARGETS)}",
    )
    args = parser.parse_args()
    configs = []
    for target in args.targets:
        if target not in TARGETS:
            raise SystemExit(f"Unknown target {target!r}. Available: {', '.join(TARGETS)}")
        configs.append(TARGETS[target])

    rows_by_target = {config.key: fetch_target_rows(config) for config in configs}
    rows_by_target = deduplicate_targets_rows(rows_by_target, configs)
    for config in configs:
        build_workbook(config, rows_by_target[config.key])


if __name__ == "__main__":
    main()
