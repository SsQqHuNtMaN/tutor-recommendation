from __future__ import annotations

import hashlib
import os
import re
import subprocess
import time
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlencode, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup

from .contact_status import STATUS_COLUMN, apply_contact_statuses
from .migrate_contact_status_column import migrate_workbook
from .dblp_research import (
    CORE_TERMS,
    RECENT_YEARS,
    RELEVANT_PATTERN,
    score_text,
    summarize_pubs,
    unique_join,
    western_name_candidates,
)
from .ranking_policy import POLICY_VERSION, evaluate_teacher, legacy_dblp_evidence
from .teacher_identity import TEACHER_ID_COLUMN, ensure_teacher_identity, teacher_record_key
from .cache_utils import configured_max_age_days, read_cached_text
from .run_manifest import (
    RunContext,
    checkpoint_fingerprint,
    context_source_rows,
    create_run_context,
    recent_years,
    write_stage_manifest,
)


SCHOOL_SLUG = os.environ.get("SCHOOL_SLUG", "sjtu")
COLLEGE_SLUG = os.environ.get("COLLEGE_SLUG", "cs")
OUTPUT_DIR = Path("outputs") / SCHOOL_SLUG / COLLEGE_SLUG
OUTPUT_PREFIX = f"{SCHOOL_SLUG}_{COLLEGE_SLUG}_teacher_match"
INPUT_PATH = OUTPUT_DIR / f"{OUTPUT_PREFIX}_dblp.xlsx"
OUTPUT_PATH = OUTPUT_DIR / f"{OUTPUT_PREFIX}_full_research.xlsx"
CHECKPOINT_PATH = OUTPUT_DIR / "full_research_checkpoint.jsonl"
ARXIV_CACHE = OUTPUT_DIR / "arxiv_cache"
WEB_CACHE = OUTPUT_DIR / "web_cache"
TODAY = date.today().isoformat()

ARXIV_API = "https://export.arxiv.org/api/query"
SOURCE_LINKS = {
    "DBLP": "https://dblp.org/",
    "arXiv": "https://export.arxiv.org/api/query",
    "教师名录": os.environ.get("FACULTY_SOURCE_URL", "https://www.cs.sjtu.edu.cn/jiaoshiml.html"),
    "PDF附件": os.environ.get("FACULTY_PDF_SOURCE_URL", ""),
}

VENUE_PATTERN = re.compile(
    r"\b(CVPR|ICCV|ECCV|NeurIPS|NIPS|ICLR|ICML|AAAI|IJCAI|ACL|EMNLP|NAACL|"
    r"CoRL|RSS|IROS|ICRA|SIGGRAPH|KDD|WWW|WSDM|CIKM|SIGIR|MM|ACM MM|"
    r"TPAMI|T-PAMI|IJCV|TRO|T-RO|TASE|T-ASE|TMM|TMI|MICCAI|MobiCom|OSDI)\b",
    re.I,
)
RECENT_YEARS = set(recent_years())
YEAR_PATTERN = re.compile(rf"\b({'|'.join(sorted(RECENT_YEARS))})\b")
DBLP_SCORE_CAP = 120
ILLEGAL_EXCEL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
WEB_SEARCH_COLUMNS = [
    "WebSearch状态",
    "WebSearch置信度",
    "WebSearch证据条数",
    "WebSearch关键词",
    "WebSearch代表证据",
    "WebSearch来源URL",
    "WebSearch建议",
]


def norm_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    value = str(value).replace("\u3000", " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", value).strip()


def clean_excel_cell(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_EXCEL_CHARS_RE.sub("", value)
    return value


def cache_path(cache_dir: Path, key: str, suffix: str) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()
    return cache_dir / f"{digest}.{suffix}"


def valid_web_payload(text: str) -> bool:
    stripped = text.lstrip()
    lowered = stripped[:500].lower()
    if not stripped:
        return False
    if "gateway time-out" in lowered or "service unavailable" in lowered:
        return False
    return True


def fetch_with_curl(url: str, cache_dir: Path, suffix: str, timeout: int = 25) -> str:
    path = cache_path(cache_dir, url, suffix)
    cached = read_cached_text(path, configured_max_age_days("RESEARCH_CACHE_MAX_AGE_DAYS", 7))
    if cached is not None:
        if valid_web_payload(cached):
            return cached
        path.unlink(missing_ok=True)

    session = requests.Session()
    session.trust_env = False
    try:
        response = session.get(
            url,
            timeout=min(timeout, 8),
            headers={"User-Agent": "Mozilla/5.0 sjtu-research-match/1.0"},
        )
        if response.status_code == 200 and valid_web_payload(response.text):
            path.write_text(response.text, encoding="utf-8")
            return response.text
    except Exception:
        pass

    command = [
        "curl.exe",
        "-L",
        "--silent",
        "--show-error",
        "--connect-timeout",
        "8",
        "--max-time",
        str(timeout),
        url,
    ]
    completed = subprocess.run(
        command,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout + 10,
    )
    if completed.returncode == 0 and valid_web_payload(completed.stdout):
        path.write_text(completed.stdout, encoding="utf-8")
        return completed.stdout
    raise RuntimeError(completed.stderr.strip() or f"curl exit {completed.returncode}")


def load_dblp_detail_sheet() -> pd.DataFrame:
    """Carry DBLP publication details from stage 2 into the final workbook."""
    if not INPUT_PATH.exists():
        return pd.DataFrame()
    for sheet_name in ("DBLP近三年论文明细", "DBLP近三年明细"):
        try:
            return pd.read_excel(INPUT_PATH, sheet_name=sheet_name)
        except ValueError:
            continue
    return pd.DataFrame()


def load_existing_web_search() -> tuple[dict[str, dict[str, Any]], pd.DataFrame]:
    if not OUTPUT_PATH.exists():
        return {}, pd.DataFrame()
    try:
        existing = pd.read_excel(OUTPUT_PATH, sheet_name="全量教师名录")
    except (OSError, ValueError):
        existing = pd.DataFrame()
    rows: dict[str, dict[str, Any]] = {}
    if not existing.empty:
        for _, row in existing.iterrows():
            row_dict = ensure_teacher_identity(SCHOOL_SLUG, COLLEGE_SLUG, row.to_dict())
            rows[checkpoint_row_key(row_dict)] = row_dict
    try:
        detail = pd.read_excel(OUTPUT_PATH, sheet_name="WebSearch证据明细")
    except (OSError, ValueError):
        detail = pd.DataFrame()
    return rows, detail


def normalize_url(url: str) -> str:
    url = norm_text(url).strip(" ;,，")
    if not url or url.lower() == "nan":
        return ""
    if not re.match(r"https?://", url, flags=re.I):
        url = "https://" + url
    return url


def split_urls(value: Any) -> list[str]:
    text = norm_text(value)
    urls = re.findall(r"https?://[^\s;，,]+|(?:[a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}[^\s;，,]*", text)
    return [url for url in (normalize_url(url) for url in urls) if url]


def parse_arxiv_entries(xml_text: str, query_name: str) -> dict[str, Any]:
    ns = {
        "a": "http://www.w3.org/2005/Atom",
        "opensearch": "http://a9.com/-/spec/opensearch/1.1/",
    }
    root = ET.fromstring(xml_text.encode("utf-8"))
    total = int(root.findtext("opensearch:totalResults", "0", namespaces=ns) or 0)
    query_key = re.sub(r"[^a-z]", "", query_name.lower())
    entries: list[dict[str, str]] = []

    for entry in root.findall("a:entry", ns):
        authors = [
            norm_text(author.findtext("a:name", default="", namespaces=ns))
            for author in entry.findall("a:author", ns)
        ]
        author_keys = {re.sub(r"[^a-z]", "", author.lower()) for author in authors}
        if query_key not in author_keys:
            continue
        published = norm_text(entry.findtext("a:published", default="", namespaces=ns))
        year = published[:4]
        if year not in RECENT_YEARS:
            continue
        title = norm_text(entry.findtext("a:title", default="", namespaces=ns)).rstrip(".")
        summary = norm_text(entry.findtext("a:summary", default="", namespaces=ns))
        link = ""
        for link_node in entry.findall("a:link", ns):
            if link_node.attrib.get("rel") == "alternate":
                link = link_node.attrib.get("href", "")
                break
        categories = [
            node.attrib.get("term", "")
            for node in entry.findall("a:category", ns)
            if node.attrib.get("term")
        ]
        entries.append(
            {
                "year": year,
                "published": published[:10],
                "title": title,
                "summary": summary,
                "categories": ", ".join(categories),
                "link": link,
                "authors": "; ".join(authors[:12]),
            }
        )
    entries.sort(key=lambda item: item["published"], reverse=True)
    return {"total": total, "entries": entries}


def query_arxiv(row: pd.Series) -> dict[str, Any]:
    if not should_deep_query(row):
        return {"状态": "未检索-低相关或同名风险", "论文": []}
    name = norm_text(row.get("姓名"))
    teacher_url = norm_text(row.get("教师主页链接"))
    candidates = western_name_candidates(name, teacher_url)[:2]
    if not candidates:
        return {"状态": "未查询-无英文姓名候选", "论文": []}

    all_entries: list[dict[str, str]] = []
    totals: list[int] = []
    errors: list[str] = []
    for candidate in candidates:
        query = f'au:"{candidate}"'
        params = urlencode(
            {
                "search_query": query,
                "start": 0,
                "max_results": 30,
                "sortBy": "submittedDate",
                "sortOrder": "descending",
            }
        )
        url = f"{ARXIV_API}?{params}"
        try:
            text = fetch_with_curl(url, ARXIV_CACHE, "xml", timeout=18)
            parsed = parse_arxiv_entries(text, candidate)
            totals.append(parsed["total"])
            all_entries.extend(parsed["entries"])
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{candidate}: {type(exc).__name__}: {exc}")
        time.sleep(0.15)

    # Deduplicate by title/link.
    dedup: dict[str, dict[str, str]] = {}
    for entry in all_entries:
        key = entry["link"] or entry["title"].lower()
        dedup[key] = entry
    entries = sorted(dedup.values(), key=lambda item: item["published"], reverse=True)

    if not entries:
        status = "无近三年arXiv命中"
        if errors and not totals:
            status = "arXiv查询失败"
        return {"状态": status, "错误": unique_join(errors), "论文": [], "总命中": max(totals or [0])}

    title_text = " ".join(f"{entry['title']} {entry['summary']} {entry['categories']}" for entry in entries)
    arxiv_score, keywords = score_text(title_text)
    dblp_author_key = re.sub(r"[^a-z]", "", norm_text(row.get("DBLP作者")).lower())
    dblp_verified = norm_text(row.get("DBLP匹配置信度")) == "高"
    dblp_titles = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", norm_text(row.get("DBLP近三年代表论文")).lower())
    author_cross_check = bool(
        dblp_author_key
        and any(
            dblp_author_key in {re.sub(r"[^a-z]", "", author.lower()) for author in entry["authors"].split("; ")}
            for entry in entries
        )
    )
    title_cross_check = bool(
        dblp_titles
        and any(
            re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", entry["title"].lower()) in dblp_titles
            for entry in entries
            if len(re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", entry["title"])) >= 12
        )
    )
    confidence = "中-DBLP交叉确认" if dblp_verified and author_cross_check and title_cross_check else "低-姓名未消歧"

    return {
        "状态": "已查询",
        "置信度": confidence,
        "总命中": max(totals or [0]),
        "论文": entries,
        "关键词": unique_join(keywords),
        "分数": arxiv_score,
        "错误": unique_join(errors),
    }


def html_to_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    meta_parts = []
    for meta in soup.find_all("meta"):
        if meta.get("name", "").lower() in {"description", "keywords", "author"} or meta.get("property", "").lower() in {
            "og:description",
            "og:title",
        }:
            meta_parts.append(norm_text(meta.get("content", "")))
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    body = soup.get_text("\n", strip=True)
    return norm_text("\n".join(meta_parts + [body]))


def extract_recent_web_snippets(text: str) -> list[str]:
    raw_lines = re.split(r"[\n\r]+|(?<=\.)\s{2,}", text)
    snippets: list[str] = []
    for line in raw_lines:
        line = norm_text(line)
        if len(line) < 25:
            continue
        if YEAR_PATTERN.search(line) or VENUE_PATTERN.search(line):
            snippets.append(line[:500])
        if len(snippets) >= 18:
            break
    return snippets


def query_web_homepage(row: pd.Series) -> dict[str, Any]:
    if not should_deep_query(row):
        return {"状态": "未检索-低相关", "证据": []}
    urls = split_urls(row.get("个人主页"))
    if not urls:
        teacher_url = norm_text(row.get("教师主页链接"))
        if teacher_url.startswith(("http://", "https://")):
            urls = [teacher_url]
    if not urls:
        return {"状态": "无个人主页URL", "证据": []}
    urls = urls[:2]
    if not urls:
        return {"状态": "无网页URL", "证据": []}

    snippets: list[str] = []
    errors: list[str] = []
    fetched_urls: list[str] = []
    all_text_parts: list[str] = []
    for url in urls:
        try:
            html = fetch_with_curl(url, WEB_CACHE, "html", timeout=15)
            text = html_to_text(html)
            fetched_urls.append(url)
            all_text_parts.append(text[:12000])
            snippets.extend(extract_recent_web_snippets(text))
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{url}: {type(exc).__name__}: {exc}")
        time.sleep(0.1)

    combined = " ".join(all_text_parts + snippets)
    web_score, keywords = score_text(combined)
    if fetched_urls:
        status = "已抓取"
    else:
        status = "网页抓取失败"
    return {
        "状态": status,
        "URL": unique_join(fetched_urls),
        "证据": snippets[:12],
        "关键词": unique_join(keywords),
        "分数": web_score,
        "错误": unique_join(errors),
    }


def should_deep_query(row: pd.Series) -> bool:
    if norm_text(row.get("是否建议套磁")) == "是":
        return True
    try:
        if float(row.get("匹配分") or 0) >= 20:
            return True
    except Exception:
        pass
    try:
        if float(row.get("DBLP近三年论文数") or 0) > 0:
            return True
    except Exception:
        pass
    if split_urls(row.get("个人主页")):
        return True
    text = " ".join(
        norm_text(row.get(col))
        for col in ["命中关键词", "研究方向", "个人简介摘要", "综合研究方向（主页+DBLP）", "DBLP近三年关键词"]
    )
    return bool(RELEVANT_PATTERN.search(text))


def prepare_web_search_columns_for_restore(df: pd.DataFrame) -> pd.DataFrame:
    for column in WEB_SEARCH_COLUMNS:
        if column == "WebSearch证据条数":
            if column not in df.columns:
                df[column] = 0
            continue
        if column not in df.columns:
            df[column] = pd.Series("", index=df.index, dtype="object")
        else:
            df[column] = df[column].astype("object")
    return df


def summarize_entries(entries: list[dict[str, str]], source: str) -> str:
    items = []
    for entry in entries[:10]:
        if source == "arxiv":
            items.append(f"[{entry['published']}] {entry['title']} ({entry['categories']}) {entry['link']}")
        else:
            items.append(norm_text(str(entry))[:300])
    return "；".join(items)


def update_recommendation(row: pd.Series, arxiv: dict[str, Any], web: dict[str, Any]) -> dict[str, Any]:
    decision = evaluate_teacher(
        row,
        dblp=legacy_dblp_evidence(row),
        arxiv=arxiv,
        web=web,
    )
    update = decision.to_columns()
    direction_bits = [
        norm_text(row.get("研究方向")),
        f"DBLP关键词：{row.get('DBLP近三年关键词')}" if norm_text(row.get("DBLP近三年关键词")) else "",
        f"arXiv关键词：{arxiv.get('关键词')}" if arxiv.get("关键词") else "",
        f"网页关键词：{web.get('关键词')}" if web.get("关键词") else "",
    ]
    update["综合研究方向（主页+DBLP+arXiv+网页）"] = unique_join(direction_bits)[:1500]
    return update


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill("solid", fgColor="D9EAD3")
    mid_fill = PatternFill("solid", fgColor="FFF2CC")
    low_fill = PatternFill("solid", fgColor="F4CCCC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx, header in enumerate(headers, start=1):
            col = get_column_letter(idx)
            header_text = norm_text(header)
            if any(token in header_text for token in ["论文", "方向", "理由", "证据"]):
                ws.column_dimensions[col].width = 58
            elif "链接" in header_text or "URL" in header_text:
                ws.column_dimensions[col].width = 36
            else:
                ws.column_dimensions[col].width = 16
        if ws.max_row >= 2 and "推荐等级" in headers:
            col = get_column_letter(headers.index("推荐等级") + 1)
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"强烈建议"'], fill=high_fill),
            )
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"可以考虑"'], fill=mid_fill),
            )
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"暂不优先"'], fill=low_fill),
            )
    wb.save(path)


def load_checkpoint() -> dict[int, dict[str, Any]]:
    if not CHECKPOINT_PATH.exists():
        return {}
    processed: dict[int, dict[str, Any]] = {}
    import json

    with CHECKPOINT_PATH.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            item = json.loads(line)
            processed[int(item["index"])] = item
    return processed


def append_checkpoint(item: dict[str, Any]) -> None:
    import json

    CHECKPOINT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with CHECKPOINT_PATH.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(item, ensure_ascii=False) + "\n")


def checkpoint_matches_current_row(row: pd.Series, item: dict[str, Any], context: RunContext) -> bool:
    return item.get("fingerprint") == checkpoint_fingerprint(row, SCHOOL_SLUG, COLLEGE_SLUG, context)


def checkpoint_row_key(row: pd.Series | dict[str, Any]) -> str:
    return teacher_record_key(SCHOOL_SLUG, COLLEGE_SLUG, row)


def checkpoint_by_teacher_key(processed: dict[int, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    keyed: dict[str, dict[str, Any]] = {}
    for item in processed.values():
        key = checkpoint_row_key(item.get("row", {}))
        if key != "::":
            keyed[key] = item
    return keyed


def checkpoint_evidence(item: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    """Rebuild arXiv/web evidence from checkpoint while using the latest DBLP row."""
    saved_row = item.get("row", {})

    arxiv_entries: list[dict[str, str]] = []
    for detail in item.get("arxiv_rows", []):
        published = norm_text(detail.get("发布日期"))
        arxiv_entries.append(
            {
                "year": published[:4],
                "published": published,
                "title": norm_text(detail.get("题名")),
                "summary": "",
                "categories": norm_text(detail.get("分类")),
                "link": norm_text(detail.get("链接")),
                "authors": norm_text(detail.get("作者")),
            }
        )
    arxiv_text = " ".join(
        [norm_text(saved_row.get("arXiv关键词")), norm_text(saved_row.get("arXiv代表论文"))]
        + [f"{entry['title']} {entry['categories']}" for entry in arxiv_entries]
    )
    arxiv_score, arxiv_keywords = score_text(arxiv_text)
    arxiv = {
        "状态": saved_row.get("arXiv状态", ""),
        "置信度": saved_row.get("arXiv置信度", ""),
        "总命中": saved_row.get("arXiv总命中", ""),
        "论文": arxiv_entries,
        "关键词": norm_text(saved_row.get("arXiv关键词")) or unique_join(arxiv_keywords),
        "分数": arxiv_score,
        "错误": saved_row.get("arXiv错误", ""),
    }

    web_snippets = [norm_text(detail.get("证据")) for detail in item.get("web_rows", [])]
    if not web_snippets and norm_text(saved_row.get("网页近三年/发表证据")):
        web_snippets = [
            norm_text(snippet)
            for snippet in str(saved_row.get("网页近三年/发表证据")).split("；")
            if norm_text(snippet)
        ]
    web_text = " ".join([norm_text(saved_row.get("网页关键词"))] + web_snippets)
    web_score, web_keywords = score_text(web_text)
    web = {
        "状态": saved_row.get("网页状态", ""),
        "URL": saved_row.get("网页URL", ""),
        "证据": web_snippets,
        "关键词": norm_text(saved_row.get("网页关键词")) or unique_join(web_keywords),
        "分数": web_score,
        "错误": saved_row.get("网页错误", ""),
    }
    return arxiv, web


def merge_current_row_with_checkpoint(row: pd.Series, item: dict[str, Any]) -> dict[str, Any]:
    arxiv, web = checkpoint_evidence(item)
    row_dict = row.to_dict()
    row_dict.update(update_recommendation(row, arxiv, web))
    row_dict.update(
        {
            "arXiv状态": arxiv.get("状态", ""),
            "arXiv置信度": arxiv.get("置信度", ""),
            "arXiv总命中": arxiv.get("总命中", ""),
            "arXiv近三年论文数": len(arxiv.get("论文", [])),
            "arXiv关键词": arxiv.get("关键词", ""),
            "arXiv代表论文": summarize_entries(arxiv.get("论文", []), "arxiv"),
            "arXiv错误": arxiv.get("错误", ""),
            "网页状态": web.get("状态", ""),
            "网页URL": web.get("URL", ""),
            "网页证据条数": len(web.get("证据", [])),
            "网页关键词": web.get("关键词", ""),
            "网页近三年/发表证据": "；".join(web.get("证据", [])),
            "网页错误": web.get("错误", ""),
        }
    )
    return row_dict


def require_complete_checkpoint_coverage(valid: int, total: int, allow_partial: bool) -> None:
    if valid != total and not allow_partial:
        raise RuntimeError(
            f"finalize-only requires 100% valid checkpoint coverage; valid={valid}/{total}. "
            "Run the normal completion stage or pass --allow-partial explicitly."
        )


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(INPUT_PATH)
    run_context = create_run_context("final", f"{SCHOOL_SLUG}_{COLLEGE_SLUG}", [INPUT_PATH])
    df = pd.read_excel(INPUT_PATH, sheet_name="全量教师名录")
    df = pd.DataFrame(
        ensure_teacher_identity(SCHOOL_SLUG, COLLEGE_SLUG, row.to_dict())
        for _, row in df.iterrows()
    )
    df = apply_contact_statuses(df, SCHOOL_SLUG, COLLEGE_SLUG)
    dblp_detail = load_dblp_detail_sheet()
    existing_web_search, web_search_detail_df = load_existing_web_search()

    raw_processed = load_checkpoint()
    checkpoint_by_key = checkpoint_by_teacher_key(raw_processed)
    processed: dict[int, dict[str, Any]] = {}
    for idx, row in df.iterrows():
        item = checkpoint_by_key.get(checkpoint_row_key(row))
        if item and checkpoint_matches_current_row(row, item, run_context):
            processed[idx] = item

    finalize_only = os.environ.get("FINALIZE_ONLY") == "1"
    allow_partial = os.environ.get("ALLOW_PARTIAL_FINALIZE") == "1"
    if finalize_only:
        require_complete_checkpoint_coverage(len(processed), len(df), allow_partial)
    for idx, row in df.iterrows():
        if idx in processed or finalize_only:
            continue
        print(f"processing {idx + 1}/{len(df)} {row.get('姓名')}", flush=True)

        arxiv = query_arxiv(row)
        web = query_web_homepage(row)
        update = update_recommendation(row, arxiv, web)

        row_dict = row.to_dict()
        row_dict.update(update)
        row_dict.update(
            {
                "arXiv状态": arxiv.get("状态", ""),
                "arXiv置信度": arxiv.get("置信度", ""),
                "arXiv总命中": arxiv.get("总命中", ""),
                "arXiv近三年论文数": len(arxiv.get("论文", [])),
                "arXiv关键词": arxiv.get("关键词", ""),
                "arXiv代表论文": summarize_entries(arxiv.get("论文", []), "arxiv"),
                "arXiv错误": arxiv.get("错误", ""),
                "网页状态": web.get("状态", ""),
                "网页URL": web.get("URL", ""),
                "网页证据条数": len(web.get("证据", [])),
                "网页关键词": web.get("关键词", ""),
                "网页近三年/发表证据": "；".join(web.get("证据", [])),
                "网页错误": web.get("错误", ""),
            }
        )
        arxiv_detail_rows: list[dict[str, Any]] = []
        for entry in arxiv.get("论文", []):
            arxiv_detail_rows.append(
                {
                    "姓名": row.get("姓名"),
                    TEACHER_ID_COLUMN: row.get(TEACHER_ID_COLUMN, ""),
                    STATUS_COLUMN: row.get(STATUS_COLUMN, ""),
                    "arXiv置信度": arxiv.get("置信度", ""),
                    "发布日期": entry["published"],
                    "题名": entry["title"],
                    "分类": entry["categories"],
                    "链接": entry["link"],
                    "作者": entry["authors"],
                }
            )
        web_detail_rows: list[dict[str, Any]] = []
        for snippet in web.get("证据", []):
            web_detail_rows.append(
                {
                    "姓名": row.get("姓名"),
                    TEACHER_ID_COLUMN: row.get(TEACHER_ID_COLUMN, ""),
                    STATUS_COLUMN: row.get(STATUS_COLUMN, ""),
                    "网页URL": web.get("URL", ""),
                    "证据": snippet,
                    "教师主页链接": row.get("教师主页链接"),
                }
            )

        checkpoint_item = {
            "index": idx,
            "fingerprint": checkpoint_fingerprint(row, SCHOOL_SLUG, COLLEGE_SLUG, run_context),
            "run_context": {
                "run_id": run_context.run_id,
                "schema_version": run_context.schema_version,
                "policy_version": run_context.policy_version,
                "profile_hash": run_context.profile_hash,
                "recent_years": run_context.recent_years,
            },
            "row": row_dict,
            "arxiv_rows": arxiv_detail_rows,
            "web_rows": web_detail_rows,
        }
        append_checkpoint(checkpoint_item)
        processed[idx] = checkpoint_item

    rows: list[dict[str, Any]] = []
    arxiv_rows: list[dict[str, Any]] = []
    web_rows: list[dict[str, Any]] = []
    for idx in range(len(df)):
        if idx not in processed:
            row = df.iloc[idx]
            row_dict = row.to_dict()
            row_dict.update(
                {
                    "综合研究方向（主页+DBLP+arXiv+网页）": norm_text(
                        row.get("综合研究方向（主页+DBLP）")
                    )
                    or norm_text(row.get("研究方向")),
                    "arXiv状态": "未深检索-时间/同名风险",
                    "arXiv置信度": "",
                    "arXiv总命中": "",
                    "arXiv近三年论文数": 0,
                    "arXiv关键词": "",
                    "arXiv代表论文": "",
                    "arXiv错误": "",
                    "网页状态": "未深检索-时间/同名风险",
                    "网页URL": "",
                    "网页证据条数": 0,
                    "网页关键词": "",
                    "网页近三年/发表证据": "",
                    "网页错误": "",
                }
            )
            rows.append(row_dict)
            continue
        item = processed[idx]
        rows.append(merge_current_row_with_checkpoint(df.iloc[idx], item))
        arxiv_rows.extend(item.get("arxiv_rows", []))
        web_rows.extend(item.get("web_rows", []))

    out = pd.DataFrame(rows)
    if existing_web_search:
        out = prepare_web_search_columns_for_restore(out)
        for idx, row in out.iterrows():
            existing = existing_web_search.get(checkpoint_row_key(row))
            if not existing:
                continue
            for column in WEB_SEARCH_COLUMNS:
                out.at[idx, column] = existing.get(column, "")
            arxiv = {
                "置信度": row.get("arXiv置信度", ""),
                "关键词": row.get("arXiv关键词", ""),
                "论文": [],
            }
            web = {
                "状态": row.get("网页状态", ""),
                "关键词": row.get("网页关键词", ""),
                "证据": [part for part in norm_text(row.get("网页近三年/发表证据")).split("；") if part],
            }
            web_search = {column: existing.get(column, "") for column in WEB_SEARCH_COLUMNS}
            policy_update = evaluate_teacher(
                row,
                dblp=legacy_dblp_evidence(row),
                arxiv=arxiv,
                web=web,
                web_search=web_search,
            ).to_columns()
            for column, value in policy_update.items():
                out.at[idx, column] = value
    out = apply_contact_statuses(out, SCHOOL_SLUG, COLLEGE_SLUG)
    dblp_detail = apply_contact_statuses(dblp_detail, SCHOOL_SLUG, COLLEGE_SLUG) if not dblp_detail.empty else dblp_detail
    arxiv_detail_df = pd.DataFrame(arxiv_rows)
    if not arxiv_detail_df.empty:
        dedup_columns = [column for column in [TEACHER_ID_COLUMN, "链接", "题名"] if column in arxiv_detail_df]
        arxiv_detail_df = arxiv_detail_df.drop_duplicates(subset=dedup_columns, keep="last")
        arxiv_detail_df = apply_contact_statuses(arxiv_detail_df, SCHOOL_SLUG, COLLEGE_SLUG)
    web_detail_df = pd.DataFrame(web_rows)
    if not web_detail_df.empty:
        dedup_columns = [column for column in [TEACHER_ID_COLUMN, "网页URL", "证据"] if column in web_detail_df]
        web_detail_df = web_detail_df.drop_duplicates(subset=dedup_columns, keep="last")
        web_detail_df = apply_contact_statuses(web_detail_df, SCHOOL_SLUG, COLLEGE_SLUG)
    if not dblp_detail.empty:
        dedup_columns = [
            column for column in [TEACHER_ID_COLUMN, "DBLP作者链接", "年份", "题名", "链接"] if column in dblp_detail
        ]
        dblp_detail = dblp_detail.drop_duplicates(subset=dedup_columns, keep="last")
    if not web_search_detail_df.empty:
        web_search_detail_df = pd.DataFrame(
            ensure_teacher_identity(SCHOOL_SLUG, COLLEGE_SLUG, row.to_dict())
            for _, row in web_search_detail_df.iterrows()
        )
        dedup_columns = [column for column in [TEACHER_ID_COLUMN, "来源URL", "标题"] if column in web_search_detail_df]
        web_search_detail_df = web_search_detail_df.drop_duplicates(subset=dedup_columns, keep="last")
        web_search_detail_df = apply_contact_statuses(web_search_detail_df, SCHOOL_SLUG, COLLEGE_SLUG)
    priority_order = {"强烈建议": 0, "可以考虑": 1, "暂不优先": 2}
    out["_order"] = out["推荐等级"].map(priority_order).fillna(9)
    out = out.sort_values(["_order", "匹配分", "名录序号"], ascending=[True, False, True]).drop(columns=["_order"])
    priority = out[out["是否建议套磁"] == "是"].copy()

    source_rows = [
        {"项目": "更新日期", "内容": TODAY},
        {"项目": "DBLP来源", "内容": SOURCE_LINKS["DBLP"]},
        {"项目": "arXiv来源", "内容": SOURCE_LINKS["arXiv"]},
        {"项目": "教师主页来源", "内容": SOURCE_LINKS["教师名录"]},
    ]
    if SOURCE_LINKS["PDF附件"]:
        source_rows.append({"项目": "PDF附件来源", "内容": SOURCE_LINKS["PDF附件"]})
    source_rows.extend(
        [
            {
                "项目": "近三年口径",
                "内容": "DBLP与arXiv按年份字段取" + "/".join(sorted(RECENT_YEARS)) + "；网页证据按相同年份窗口或顶会/期刊关键词抽取。",
            },
            {
                "项目": "消歧说明",
                "内容": "DBLP高置信优先；arXiv只有作者名，常见姓名标为低置信并仅作辅助；个人/实验室主页作为web证据补充方向和发表线索。",
            },
        ]
    )
    source_rows.extend(context_source_rows(run_context))
    source = pd.DataFrame(source_rows)
    out = out.map(clean_excel_cell)
    priority = priority.map(clean_excel_cell)
    dblp_detail = dblp_detail.map(clean_excel_cell) if not dblp_detail.empty else dblp_detail
    arxiv_detail_df = arxiv_detail_df.map(clean_excel_cell) if not arxiv_detail_df.empty else arxiv_detail_df
    web_detail_df = web_detail_df.map(clean_excel_cell) if not web_detail_df.empty else web_detail_df
    source = source.map(clean_excel_cell)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        priority.to_excel(writer, sheet_name="优先套磁名单", index=False)
        out.to_excel(writer, sheet_name="全量教师名录", index=False)
        dblp_detail.to_excel(writer, sheet_name="DBLP近三年明细", index=False)
        arxiv_detail_df.to_excel(writer, sheet_name="arXiv近三年明细", index=False)
        web_detail_df.to_excel(writer, sheet_name="网页证据明细", index=False)
        if not web_search_detail_df.empty:
            web_search_detail_df.to_excel(writer, sheet_name="WebSearch证据明细", index=False)
        source.to_excel(writer, sheet_name="匹配依据", index=False)

    style_workbook(OUTPUT_PATH)
    migrate_workbook(OUTPUT_PATH, {})
    write_stage_manifest(OUTPUT_DIR, run_context)
    print(f"rows={len(out)}")
    print(f"priority={len(priority)}")
    print(f"dblp_entries={len(dblp_detail)}")
    print(f"arxiv_entries={len(arxiv_rows)}")
    print(f"web_evidence={len(web_rows)}")
    print(f"output={OUTPUT_PATH.resolve()}")
    print(out["推荐等级"].value_counts().to_string())
    print(
        priority[
            [
                "姓名",
                "职称",
                "推荐等级",
                "匹配分",
                "DBLP近三年论文数",
                "arXiv近三年论文数",
                "网页证据条数",
                "命中关键词",
                "教师主页链接",
            ]
        ]
        .head(30)
        .to_string(index=False)
    )


if __name__ == "__main__":
    main()
