from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from .contact_status import (
    CONTACT_COLUMNS,
    CONTACT_DATE_COLUMN,
    LEGACY_CONTACT_NOTE_COLUMNS,
    CONTACT_NOTE_COLUMN,
    CONTACT_RESPONSE_COLUMN,
    CONTACT_STATUS_PATH,
    STATUS_COLUMN,
    VALID_CONTACT_STATUSES,
    contact_entry_from_store,
    has_status_entry,
    join_responses,
    load_status_store,
    normalize_status,
    row_key,
    split_responses,
    unique_join_text,
)


def target_from_path(path: Path) -> tuple[str, str]:
    parts = path.parts
    try:
        outputs_index = parts.index("outputs")
        return parts[outputs_index + 1], parts[outputs_index + 2]
    except (ValueError, IndexError):
        return "", ""


def header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(ws[1], start=1)
        if cell.value is not None and str(cell.value).strip()
    }


def copy_column_style(ws, source_col: int, target_col: int) -> None:
    ws.column_dimensions[ws.cell(row=1, column=target_col).column_letter].width = 12
    for row_idx in range(1, ws.max_row + 1):
        source = ws.cell(row=row_idx, column=source_col)
        target = ws.cell(row=row_idx, column=target_col)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def add_validation(ws, status_col: int) -> None:
    formula = '"' + ",".join(VALID_CONTACT_STATUSES) + '"'
    ws.data_validations.dataValidation = [
        dv
        for dv in ws.data_validations.dataValidation
        if not (dv.type == "list" and (dv.formula1 == formula or dv.promptTitle == STATUS_COLUMN))
    ]
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "只能选择：已套磁、先不考虑、不可能、不匹配，或留空。"
    dv.errorTitle = "无效套磁情况"
    dv.prompt = "请选择套磁情况；尚未套磁时保持空白。"
    dv.promptTitle = STATUS_COLUMN
    ws.add_data_validation(dv)
    col_letter = ws.cell(row=1, column=status_col).column_letter
    max_row = max(ws.max_row, 1000)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def row_identity(ws, row_idx: int, headers: dict[str, int]) -> dict[str, object]:
    return {
        "姓名": ws.cell(row=row_idx, column=headers.get("姓名", 0)).value,
        "教师主页链接": ws.cell(row=row_idx, column=headers.get("教师主页链接", 0)).value
        if headers.get("教师主页链接")
        else "",
        "个人主页": ws.cell(row=row_idx, column=headers.get("个人主页", 0)).value
        if headers.get("个人主页")
        else "",
    }


def cell_text(ws, row_idx: int, headers: dict[str, int], column_name: str) -> str:
    col_idx = headers.get(column_name)
    if not col_idx:
        return ""
    value = ws.cell(row=row_idx, column=col_idx).value
    return "" if value is None else str(value).strip()


def contact_entry_from_sheet(ws, row_idx: int, headers: dict[str, int]) -> dict[str, object]:
    entry: dict[str, object] = {}
    status = normalize_status(cell_text(ws, row_idx, headers, STATUS_COLUMN))
    contacted_at = cell_text(ws, row_idx, headers, CONTACT_DATE_COLUMN)
    responses = split_responses(cell_text(ws, row_idx, headers, CONTACT_RESPONSE_COLUMN))
    note = unique_join_text(
        [
            cell_text(ws, row_idx, headers, CONTACT_NOTE_COLUMN),
            *(cell_text(ws, row_idx, headers, column) for column in LEGACY_CONTACT_NOTE_COLUMNS),
        ]
    )
    if status:
        entry["status"] = status
    if contacted_at:
        entry["contacted_at"] = contacted_at
    if responses:
        entry["responses"] = responses
    if note:
        entry["note"] = note
    return entry


def update_contact_values(
    ws,
    school_slug: str,
    college_slug: str,
    store: dict,
    authoritative: bool,
    existing_by_key: dict[str, dict[str, object]] | None = None,
) -> None:
    headers = header_map(ws)
    for row_idx in range(2, ws.max_row + 1):
        row = row_identity(ws, row_idx, headers)
        key = row_key(school_slug, college_slug, row)
        stored_entry = contact_entry_from_store(store, key)
        current_entry = contact_entry_from_sheet(ws, row_idx, headers) or (existing_by_key or {}).get(key, {})
        if authoritative:
            entry = stored_entry if has_status_entry(store, key) else {}
        else:
            entry = {**current_entry, **stored_entry}
        ws.cell(row=row_idx, column=headers[STATUS_COLUMN]).value = normalize_status(entry.get("status"))
        ws.cell(row=row_idx, column=headers[CONTACT_DATE_COLUMN]).value = str(entry.get("contacted_at") or "")
        ws.cell(row=row_idx, column=headers[CONTACT_RESPONSE_COLUMN]).value = join_responses(entry.get("responses", []))
        ws.cell(row=row_idx, column=headers[CONTACT_NOTE_COLUMN]).value = str(entry.get("note") or "")


def contact_columns_are_ordered(headers: dict[str, int]) -> bool:
    if "姓名" not in headers:
        return False
    if any(column in headers for column in LEGACY_CONTACT_NOTE_COLUMNS):
        return False
    start = headers["姓名"] + 1
    return all(headers.get(column) == start + offset for offset, column in enumerate(CONTACT_COLUMNS))


def migrate_sheet(ws, school_slug: str, college_slug: str, store: dict, authoritative: bool = False) -> bool:
    headers = header_map(ws)
    if "姓名" not in headers:
        return False

    existing_by_key: dict[str, dict[str, object]] = {}
    for row_idx in range(2, ws.max_row + 1):
        key = row_key(school_slug, college_slug, row_identity(ws, row_idx, headers))
        entry = contact_entry_from_sheet(ws, row_idx, headers)
        if entry:
            existing_by_key[key] = entry
    if contact_columns_are_ordered(headers):
        update_contact_values(ws, school_slug, college_slug, store, authoritative, existing_by_key=existing_by_key)
        add_validation(ws, headers[STATUS_COLUMN])
        return True

    contact_like_columns = CONTACT_COLUMNS + LEGACY_CONTACT_NOTE_COLUMNS
    for column in sorted((headers[column] for column in contact_like_columns if column in headers), reverse=True):
        ws.delete_cols(column, 1)

    headers = header_map(ws)
    name_col = headers["姓名"]
    insert_col = name_col + 1
    ws.insert_cols(insert_col, len(CONTACT_COLUMNS))
    for offset, column_name in enumerate(CONTACT_COLUMNS):
        target_col = insert_col + offset
        copy_column_style(ws, name_col, target_col)
        ws.cell(row=1, column=target_col).value = column_name

    update_contact_values(ws, school_slug, college_slug, store, authoritative, existing_by_key=existing_by_key)
    add_validation(ws, insert_col)
    if ws.auto_filter and ws.auto_filter.ref:
        ws.auto_filter.ref = ws.dimensions
    return True


def migrate_workbook(path: Path, store: dict, authoritative: bool = False) -> bool:
    school_slug, college_slug = target_from_path(path)
    wb = load_workbook(path)
    changed = False
    for ws in wb.worksheets:
        changed = migrate_sheet(ws, school_slug, college_slug, store, authoritative=authoritative) or changed
    if changed:
        wb.save(path)
    return changed


def iter_current_workbooks() -> list[Path]:
    paths = []
    for path in Path("outputs").rglob("*.xlsx"):
        if "archive" in path.parts:
            continue
        paths.append(path)
    return sorted(paths)


def main() -> None:
    store = load_status_store(CONTACT_STATUS_PATH)
    changed_paths = []
    failed_paths = []
    for path in iter_current_workbooks():
        try:
            if migrate_workbook(path, store):
                changed_paths.append(path)
                print(f"updated {path}")
            else:
                print(f"ok {path}")
        except Exception as exc:  # noqa: BLE001
            failed_paths.append(path)
            print(f"failed {path}: {type(exc).__name__}: {exc}")
    print(f"workbooks={len(iter_current_workbooks())}")
    print(f"changed={len(changed_paths)}")
    print(f"failed={len(failed_paths)}")


if __name__ == "__main__":
    main()
